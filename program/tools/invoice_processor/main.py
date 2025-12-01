"""
开票信息处理工具
从 Excel 文件生成开票 TXT 格式
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from pathlib import Path
import os
from datetime import datetime
import sys

# 添加父目录到路径以导入theme模块
sys.path.insert(0, str(Path(__file__).parent.parent.parent))
from theme import get_colors
from razer_ui import Razer3DCard
from unified_button import UnifiedButton
from theme_toggle import ThemeToggleButton


class InvoiceProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Workit - 开票信息处理工具")
        self.root.geometry("900x600")
        
        # 不设置窗口图标（用户不需要）
        
        # Razer风格配色 - 自动跟随系统深色/浅色模式
        self.colors = get_colors()
        
        self.root.configure(bg=self.colors['bg_main'])
        self.excel_file = None
        
        self.create_widgets()
        self.center_window()
    
    def center_window(self):
        """窗口居中"""
        self.root.update_idletasks()
        width = 900
        height = 600
        x = (self.root.winfo_screenwidth() // 2) - (width // 2)
        y = (self.root.winfo_screenheight() // 2) - (height // 2)
        self.root.geometry(f'{width}x{height}+{x}+{y}')
    
    def create_widgets(self):
        """创建界面组件"""
        # 顶部标题区域
        header_frame = tk.Frame(self.root, bg=self.colors['bg_main'], height=80)
        header_frame.pack(fill=tk.X)
        header_frame.pack_propagate(False)
        
        # 统一使用相对坐标，让四个控件都在 header 垂直居中（整体略微下移）
        # 左侧返回按钮
        back_btn = tk.Label(
            header_frame,
            text="< 返回首页",
            font=("Microsoft YaHei UI", 10),
            bg=self.colors['bg_main'],
            fg=self.colors['text_muted'],
            cursor="hand2"
        )
        back_btn.place(relx=0.0, rely=0.61, x=40, anchor='w')
        back_btn.bind("<Button-1>", lambda e: self.back_to_launcher())
        back_btn.bind("<Enter>", lambda e: back_btn.config(fg=self.colors['text_primary']))
        back_btn.bind("<Leave>", lambda e: back_btn.config(fg=self.colors['text_muted']))
        
        # 右侧主题切换按钮（自定义日月图标）
        theme_btn = ThemeToggleButton(header_frame, command=self.toggle_theme)
        # 稍微上移一点点，让圆形图标看起来与文字居中
        theme_btn.place(relx=1.0, rely=0.58, x=-40, anchor='e')

        # 右上角帮助按钮（?）
        help_btn = tk.Label(
            header_frame,
            text="?",
            font=("Microsoft YaHei UI", 13, "bold"),
            bg=self.colors['bg_main'],
            fg=self.colors['text_muted'],
            cursor="hand2"
        )
        help_btn.place(relx=1.0, rely=0.61, x=-80, anchor='e')
        help_btn.bind("<Button-1>", lambda e: self.open_help())
        help_btn.bind("<Enter>", lambda e: help_btn.config(fg=self.colors['text_primary']))
        help_btn.bind("<Leave>", lambda e: help_btn.config(fg=self.colors['text_muted']))
        
        # 中间标题（真正垂直居中）
        title_container = tk.Frame(header_frame, bg=self.colors['bg_main'])
        title_container.place(relx=0.5, rely=0.61, anchor='center')
        
        title_label = tk.Label(
            title_container,
            text="开票信息处理工具",
            font=("Microsoft YaHei UI", 24, "bold"),
            bg=self.colors['bg_main'],
            fg=self.colors['text_primary']
        )
        title_label.pack()
        
        # 主内容区域
        content_frame = tk.Frame(self.root, bg=self.colors['bg_main'])
        content_frame.pack(fill=tk.BOTH, expand=True, padx=35, pady=35)
        
        # 第一张卡片：选择 Excel 文件
        file_card_container, file_card = self.create_card(content_frame)
        file_card_container.pack(fill=tk.X, pady=(0, 18))
        self.create_file_section(file_card)

        # 第二张卡片：处理选项
        options_card_container, options_card = self.create_card(content_frame)
        options_card_container.pack(fill=tk.X, pady=(0, 18))
        self.create_options_section(options_card)
    
    def create_card(self, parent):
        """创建Razer 3D拟物化卡片"""
        card_3d = Razer3DCard(parent)
        content = card_3d.get_content()
        
        content_padded = tk.Frame(content, bg=self.colors['bg_card'])
        content_padded.pack(fill=tk.BOTH, expand=True, padx=30, pady=22)
        
        return card_3d, content_padded
    
    def create_file_section(self, parent):
        """创建文件选择区域"""
        tk.Label(
            parent,
            text="选择 Excel 文件",
            font=("Microsoft YaHei UI", 14, "bold"),
            bg=self.colors['bg_card'],
            fg=self.colors['text_primary']
        ).pack(anchor='w', pady=(0, 15))
        
        # 文件状态显示
        file_status_frame = tk.Frame(parent, bg=self.colors['bg_main'])
        file_status_frame.pack(fill=tk.X, pady=(0, 15))
        
        self.file_label = tk.Label(
            file_status_frame,
            text="尚未选择文件",
            font=("Microsoft YaHei UI", 10),
            bg=self.colors['bg_main'],
            fg=self.colors['text_muted'],
            anchor=tk.W,
            padx=15,
            pady=12
        )
        self.file_label.pack(fill=tk.X)

        # 第一行按钮：选择文件 + 清除
        button_row = tk.Frame(parent, bg=self.colors['bg_card'])
        button_row.pack(fill=tk.X)

        select_btn = UnifiedButton(
            button_row,
            text="选择 Excel 文件",
            command=self.select_file,
            style="primary",
            width=140,
            height=40
        )
        select_btn.pack(side=tk.LEFT, padx=(0, 8))

        self.clear_btn = UnifiedButton(
            button_row,
            text="清除",
            command=self.clear_file,
            style="secondary",
            width=80,
            height=40,
            state="disabled"
        )
        self.clear_btn.pack(side=tk.RIGHT)

    def create_options_section(self, parent):
        """创建处理选项卡片内容"""
        tk.Label(
            parent,
            text="处理选项",
            font=("Microsoft YaHei UI", 14, "bold"),
            bg=self.colors['bg_card'],
            fg=self.colors['text_primary']
        ).pack(anchor='w', pady=(0, 15))

        # 主容器
        options_frame = tk.Frame(parent, bg=self.colors['bg_card'])
        options_frame.pack(fill=tk.X)

        # 第一行：唯品会
        vip_row = tk.Frame(options_frame, bg=self.colors['bg_card'])
        vip_row.pack(fill=tk.X, pady=(0, 8))

        tk.Label(
            vip_row,
            text="唯品会：",
            font=("Microsoft YaHei UI", 10),
            bg=self.colors['bg_card'],
            fg=self.colors['text_primary']
        ).pack(side=tk.LEFT, padx=(0, 10))

        self.vip_btn = UnifiedButton(
            vip_row,
            text="处理发票",
            command=self.process_vip_invoice,
            style="primary",
            height=32,
            auto_width=True,
            state="disabled"
        )
        self.vip_btn.pack(side=tk.LEFT, padx=(0, 6))

        self.vip_folder_btn = UnifiedButton(
            vip_row,
            text="整理发票文件夹",
            command=self.process_vip_folder,
            style="secondary",
            height=32,
            state="disabled"
        )
        self.vip_folder_btn.pack(side=tk.LEFT, padx=(0, 6))

        # 第二行：小米有品
        xiaomi_row = tk.Frame(options_frame, bg=self.colors['bg_card'])
        xiaomi_row.pack(fill=tk.X)

        tk.Label(
            xiaomi_row,
            text="小米有品：",
            font=("Microsoft YaHei UI", 10),
            bg=self.colors['bg_card'],
            fg=self.colors['text_primary']
        ).pack(side=tk.LEFT, padx=(0, 10))

        self.xiaomi_btn = UnifiedButton(
            xiaomi_row,
            text="处理发票",
            command=self.process_xiaomi_invoice,
            style="primary",
            height=32,
            auto_width=True,
            state="disabled"
        )
        self.xiaomi_btn.pack(side=tk.LEFT, padx=(0, 6))

    def open_help(self):
        """显示使用说明：使用内置的轻量级 Markdown 渲染（纯 Tkinter 实现，无外部依赖）。"""
        try:
            from tkinter import Canvas, Frame

            doc_path = Path(__file__).parent / "开票信息处理工具使用说明.md"
            if not doc_path.exists():
                messagebox.showinfo("提示", f"未找到使用说明文件：\n{doc_path}")
                return

            raw = doc_path.read_text(encoding="utf-8", errors="ignore")

            # 创建窗口
            help_win = tk.Toplevel(self.root)
            help_win.title("开票信息处理工具 - 使用说明")
            help_win.configure(bg=self.colors['bg_main'])
            # 居中
            help_win.update_idletasks()
            w, h = 800, 600
            sw = help_win.winfo_screenwidth()
            sh = help_win.winfo_screenheight()
            x = (sw // 2) - (w // 2)
            y = (sh // 2) - (h // 2)
            help_win.geometry(f"{w}x{h}+{x}+{y}")
            help_win.transient(self.root)
            help_win.grab_set()

            # 滚动容器：Canvas + 内部 Frame
            canvas = Canvas(help_win, bg=self.colors['bg_main'], highlightthickness=0)
            scrollbar = tk.Scrollbar(help_win, orient="vertical", command=canvas.yview)
            canvas.configure(yscrollcommand=scrollbar.set)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

            inner = Frame(canvas, bg=self.colors['bg_main'])
            canvas.create_window((0, 0), window=inner, anchor='nw')

            def _on_configure(event):
                canvas.configure(scrollregion=canvas.bbox("all"))

            inner.bind("<Configure>", _on_configure)

            # 绑定鼠标滚轮，实现滚动
            def _on_mousewheel(event):
                # Windows 上使用 event.delta，每次滚动一般为 120 的倍数
                if event.delta:
                    canvas.yview_scroll(int(-event.delta / 120), "units")

            canvas.bind_all("<MouseWheel>", _on_mousewheel)

            # 简单的 Markdown 解析：按行处理
            import re
            lines = raw.splitlines()
            for line in lines:
                stripped = line.rstrip()

                # 分隔线 --- 视为一条空白间距
                if stripped.strip() == "---":
                    tk.Label(inner, text="", bg=self.colors['bg_main']).pack(pady=2)
                    continue

                # 空行 → 加一点垂直间距
                if not stripped.strip():
                    tk.Label(inner, text="", bg=self.colors['bg_main']).pack()
                    continue

                # 标题行
                m = re.match(r"^\s*(#{1,6})\s+(.*)$", stripped)
                if m:
                    level = len(m.group(1))
                    text = m.group(2)
                    # 去掉粗体和反引号
                    text = text.replace("**", "").replace("`", "")
                    size = 16 if level <= 2 else 12
                    weight = "bold"
                    tk.Label(
                        inner,
                        text=text,
                        font=("Microsoft YaHei UI", size, weight),
                        bg=self.colors['bg_main'],
                        fg=self.colors['text_primary'],
                        anchor='w',
                        justify='left',
                        wraplength=760
                    ).pack(fill=tk.X, padx=12, pady=(6 if level <= 2 else 4, 2))
                    continue

                # 无序列表项
                m = re.match(r"^\s*[-*+]\s+(.*)$", stripped)
                if m:
                    text = m.group(1).replace("**", "").replace("`", "")
                    text = "• " + text
                    tk.Label(
                        inner,
                        text=text,
                        font=("Microsoft YaHei UI", 10),
                        bg=self.colors['bg_main'],
                        fg=self.colors['text_primary'],
                        anchor='w',
                        justify='left',
                        wraplength=760
                    ).pack(fill=tk.X, padx=24, pady=1)
                    continue

                # 普通段落（去掉粗体标记 ** 和反引号 `）
                text = stripped.replace("**", "").replace("`", "")
                tk.Label(
                    inner,
                    text=text,
                    font=("Microsoft YaHei UI", 10),
                    bg=self.colors['bg_main'],
                    fg=self.colors['text_primary'],
                    anchor='w',
                    justify='left',
                    wraplength=760
                ).pack(fill=tk.X, padx=12, pady=1)

        except Exception as e:
            messagebox.showerror("错误", f"无法打开使用说明：{e}")
    
    def create_preview_section(self, parent):
        """创建预览区域"""
        tk.Label(
            parent,
            text="开票信息预览",
            font=("Microsoft YaHei UI", 14, "bold"),
            bg=self.colors['bg_card'],
            fg=self.colors['text_primary']
        ).pack(anchor='w', pady=(0, 15))
        
        # 预览文本框
        preview_frame = tk.Frame(parent, bg=self.colors['bg_main'])
        preview_frame.pack(fill=tk.BOTH, expand=True)
        
        scrollbar = tk.Scrollbar(preview_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.preview_text = tk.Text(
            preview_frame,
            font=("Microsoft YaHei UI", 10),
            bg=self.colors['bg_main'],
            fg=self.colors['text_primary'],
            yscrollcommand=scrollbar.set,
            relief=tk.FLAT,
            borderwidth=0,
            wrap=tk.WORD,
            state=tk.DISABLED
        )
        self.preview_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.config(command=self.preview_text.yview)
    
    def select_file(self):
        """选择 Excel 文件"""
        file_path = filedialog.askopenfilename(
            title="选择 Excel 文件",
            filetypes=[
                ("Excel 文件", "*.xlsx"),
                ("所有文件", "*.*")
            ]
        )
        
        if file_path:
            self.excel_file = file_path
            self.file_label.config(
                text=f"✓ {Path(file_path).name}",
                fg=self.colors['text_primary'],
                font=("Microsoft YaHei UI", 10, "bold")
            )
            
            # 启用按钮
            self.vip_btn.config_state("normal")
            self.xiaomi_btn.config_state("normal")
            self.vip_folder_btn.config_state("normal")
            self.clear_btn.config_state("normal")
            
            # 更新预览提示
            self.update_preview()
    
    def clear_file(self):
        """清除文件"""
        self.excel_file = None
        self.file_label.config(
            text="尚未选择文件",
            fg=self.colors['text_muted'],
            font=("Microsoft YaHei UI", 10)
        )
        
        # 禁用按钮
        self.vip_btn.config_state("disabled")
        self.xiaomi_btn.config_state("disabled")
        self.vip_folder_btn.config_state("disabled")
        self.clear_btn.config_state("disabled")
        # 预览区域已移除，不再需要清空文本
    
    def update_preview(self):
        """简单打印当前选择状态（无UI预览）。"""
        if not self.excel_file:
            print("[开票工具] 尚未选择 Excel 文件")
        else:
            print(f"[开票工具] 已选择文件: {Path(self.excel_file).name}")

    def log_to_preview(self, message, clear=False):
        """将日志输出到控制台（预览UI已移除）。"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        print(f"[{timestamp}] {message}")

    def _load_worksheet(self, xlsx_path):
        """加载工作表，优先第一个工作表"""
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        return wb, ws

    def process_vip_invoice(self):
        """处理唯品会开票：表1 → 表2"""
        if not self.excel_file:
            messagebox.showwarning("提示", "请先选择唯品会原始 Excel 文件（表1）")
            return
        
        try:
            xlsx_path = Path(self.excel_file)
            self.log_to_preview(f"开始处理唯品会文件：{xlsx_path}", clear=True)
            wb, ws = self._load_worksheet(xlsx_path)
            
            # 读取表头映射
            header_map = {}
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=1, column=col).value
                if v is not None:
                    header_map[str(v).strip()] = col
            
            def col(name):
                return header_map.get(name)
            
            required_cols = [
                "订单号", "发票状态", "发票抬头", "纳税人识别号", "发票金额",
                "红冲状态", "开票内容明细", "是否红冲"
            ]
            missing = [c for c in required_cols if c not in header_map]
            if missing:
                raise ValueError(f"唯品会表头缺少列：{', '.join(missing)}")
            
            rows_out = []
            from datetime import date
            today = date.today()
            date_str = f"{today.month}月{today.day}日"
            
            import re
            for r in range(2, ws.max_row + 1):
                status = ws.cell(row=r, column=col("发票状态")).value
                if str(status).strip() != "未开具":
                    continue
                
                order_no = ws.cell(row=r, column=col("订单号")).value
                title = ws.cell(row=r, column=col("发票抬头")).value or ""
                tax_id = ws.cell(row=r, column=col("纳税人识别号")).value or ""
                amount = ws.cell(row=r, column=col("发票金额")).value or 0
                rc_status = ws.cell(row=r, column=col("红冲状态")).value
                is_red = ws.cell(row=r, column=col("是否红冲")).value
                detail = ws.cell(row=r, column=col("开票内容明细")).value or ""
                
                # 发票类型
                invoice_type = "个人"
                invoice_no = ""  # 发票号（冲红需填写）
                rc_text = str(rc_status).strip() if rc_status is not None else ""
                is_red_text = str(is_red).strip() if is_red is not None else ""

                # 仅当明确是红冲时才判定为“冲红”
                # 条件：
                # 1) “是否红冲”列为“是”，或
                # 2) “红冲状态”文本中包含“红冲”，但不包含“非红冲”
                is_red_invoice = False
                if (is_red_text == "是") or ("红冲" in rc_text and "非红冲" not in rc_text):
                    is_red_invoice = True
                elif str(tax_id).strip():
                    # 有税号且不是红冲，视为普票
                    invoice_type = "普票"
                
                # 数量
                qty = 1
                m = re.search(r"\*(\d+)", str(detail))
                if m:
                    try:
                        qty = int(m.group(1))
                    except ValueError:
                        qty = 1
                
                try:
                    qty_f = float(qty) if qty else 1.0
                except Exception:
                    qty_f = 1.0
                amt_f = float(amount) if amount not in (None, "") else 0.0
                unit_price = amt_f / qty_f if qty_f else amt_f
                unit_price = round(unit_price, 2)
                
                # 红冲时单价和金额改为负数
                if is_red_invoice:
                    unit_price = -unit_price
                    amt_f = -amt_f
                
                row = [
                    invoice_type,          # 发票类型（必须填写）
                    "",                    # 发票号（冲红需填写）
                    date_str,              # 开票日期
                    str(order_no or ""),  # 订单号
                    str(title or ""),     # 客户名称
                    str(tax_id or ""),    # 客户税号
                    "",                   # 客户地址
                    "",                   # 客户电话
                    "",                   # 开户银行
                    "",                   # 银行账号
                    "无人机",             # 货物名称
                    "",                   # 规格型号
                    "台",                 # 单位
                    qty,                   # 数量
                    unit_price,            # 单价（含税）- 红冲时为负数
                    "",                   # 税率
                    amt_f,                 # 金额（含税）- 红冲时为负数
                    "红冲" if is_red_invoice else "",  # 备注（红冲时写"红冲"）
                ]
                rows_out.append(row)
            
            if not rows_out:
                messagebox.showwarning("提示", "没有找到发票状态为‘未开具’的记录")
                self.log_to_preview("没有找到未开具记录，未生成文件。")
                return
            
            # 写出到新的 Excel
            from openpyxl import Workbook
            out_wb = Workbook()
            out_ws = out_wb.active
            headers = [
                "发票类型（必须填写）",
                "发票号（冲红需填写）",
                "开票日期",
                "订单号",
                "客户名称（必须填写）",
                "客户税号（必须填写 公司必填）",
                "客户地址（选填）",
                "客户电话（选填）",
                "开户银行（选填）",
                "银行账号（选填）",
                "货物（劳务）名称（必须填写）",
                "规格型号（选填）",
                "单位（必须填写）",
                "数量（必须填写）",
                "单价（含税）（选填）",
                "税率",
                "金额（含税）（必须填写）",
                "备注",
            ]
            out_ws.append(headers)
            for row in rows_out:
                out_ws.append(row)

            # 在第二个工作表中附带原始数据，方便对照
            orig_sheet = out_wb.create_sheet("原始数据")
            for src_row in ws.iter_rows(values_only=True):
                orig_sheet.append(list(src_row))
            
            # 生成输出路径，若已存在则追加序号避免覆盖和权限问题
            base_name = xlsx_path.stem + "_唯品会开票"
            out_path = xlsx_path.with_name(base_name + ".xlsx")
            if out_path.exists():
                for idx in range(2, 100):
                    candidate = xlsx_path.with_name(f"{base_name}_{idx}.xlsx")
                    if not candidate.exists():
                        out_path = candidate
                        break
            out_wb.save(out_path)
            self.log_to_preview(f"唯品会处理完成，共导出 {len(rows_out)} 条记录 → {out_path}")
            messagebox.showinfo("成功", f"唯品会开票文件已生成：\n{out_path}")
        except ImportError:
            messagebox.showerror("错误", "openpyxl 模块未安装！\n\n请先安装：pip install openpyxl")
        except Exception as e:
            self.log_to_preview(f"唯品会处理失败：{e}")
            messagebox.showerror("错误", f"唯品会处理失败：{e}")

    def process_xiaomi_invoice(self):
        """处理小米有品开票：表3 → 表4"""
        if not self.excel_file:
            messagebox.showwarning("提示", "请先选择小米有品原始 Excel 文件（表3）")
            return
        
        try:
            xlsx_path = Path(self.excel_file)
            self.log_to_preview(f"开始处理小米有品文件：{xlsx_path}", clear=True)
            wb, ws = self._load_worksheet(xlsx_path)
            
            # 表头映射
            header_map = {}
            for col in range(1, ws.max_column + 1):
                v = ws.cell(row=1, column=col).value
                if v is not None:
                    header_map[str(v).strip()] = col
            
            def col(name):
                return header_map.get(name)
            
            required_cols = [
                "订单号", "开票状态", "金额(包含运费)(元)", "商品数量",
                "开票类型", "发票抬头", "纳税人识别号",
                "开户银行", "银行卡号", "企业电话", "企业地址",
            ]
            missing = [c for c in required_cols if c not in header_map]
            if missing:
                raise ValueError(f"小米有品表头缺少列：{', '.join(missing)}")
            
            rows_out = []
            from datetime import date
            today = date.today()
            date_str = f"{today.year}/{today.month}/{today.day}"
            
            for r in range(2, ws.max_row + 1):
                status = ws.cell(row=r, column=col("开票状态")).value
                if str(status).strip() != "待开票":
                    continue
                
                order_no = ws.cell(row=r, column=col("订单号")).value
                amount = ws.cell(row=r, column=col("金额(包含运费)(元)")).value or 0
                qty = ws.cell(row=r, column=col("商品数量")).value or 1
                invoice_type_raw = ws.cell(row=r, column=col("开票类型")).value or ""
                title = ws.cell(row=r, column=col("发票抬头")).value or ""
                tax_id = ws.cell(row=r, column=col("纳税人识别号")).value or ""
                bank = ws.cell(row=r, column=col("开户银行")).value or ""
                bank_no = ws.cell(row=r, column=col("银行卡号")).value or ""
                phone = ws.cell(row=r, column=col("企业电话")).value or ""
                addr = ws.cell(row=r, column=col("企业地址")).value or ""
                
                # 发票类型映射
                it = str(invoice_type_raw).strip()
                if it == "个人电子发票":
                    invoice_type = "个人发票"
                elif it == "企业电子发票":
                    invoice_type = "电子普通发票"
                else:
                    invoice_type = it or "个人发票"
                
                try:
                    qty_f = float(qty) if qty else 1.0
                except Exception:
                    qty_f = 1.0
                amt_f = float(amount) if amount not in (None, "") else 0.0
                unit_price = amt_f / qty_f if qty_f else amt_f
                unit_price = round(unit_price, 2)
                
                row = [
                    invoice_type,          # 发票类型
                    date_str,               # 开票日期
                    str(order_no or ""),  # 订单号
                    str(title or ""),     # 客户名称
                    str(tax_id or ""),    # 客户税号
                    str(addr or ""),      # 客户地址
                    str(phone or ""),     # 客户电话
                    str(bank or ""),      # 开户银行
                    str(bank_no or ""),   # 银行账号
                    "无人机",             # 货物名称
                    "",                   # 规格型号
                    "台",                 # 单位
                    qty,                   # 数量
                    unit_price,            # 单价（含税）
                    "",                   # 税率
                    amt_f,                 # 金额（含税）
                    "",                   # 备注
                ]
                rows_out.append(row)
            
            if not rows_out:
                messagebox.showwarning("提示", "没有找到开票状态为‘待开票’的记录")
                self.log_to_preview("没有找到待开票记录，未生成文件。")
                return
            
            from openpyxl import Workbook
            out_wb = Workbook()
            out_ws = out_wb.active
            headers = [
                "发票类型（必须填写）",
                "开票日期",
                "订单号",
                "客户名称（必须填写）",
                "客户税号（必须填写）",
                "客户地址（选填）",
                "客户电话（选填）",
                "开户银行（选填）",
                "银行账号（选填）",
                "货物（劳务）名称（必须填写）",
                "规格型号（选填）",
                "单位（必须填写）",
                "数量（必须填写）",
                "单价（含税）（选填）",
                "税率",
                "金额（含税）（必须填写）",
                "备注",
            ]
            out_ws.append(headers)
            for row in rows_out:
                out_ws.append(row)

            # 在第二个工作表中附带原始数据，方便对照
            orig_sheet = out_wb.create_sheet("原始数据")
            for src_row in ws.iter_rows(values_only=True):
                orig_sheet.append(list(src_row))
            
            base_name = xlsx_path.stem + "_小米有品开票"
            out_path = xlsx_path.with_name(base_name + ".xlsx")
            if out_path.exists():
                for idx in range(2, 100):
                    candidate = xlsx_path.with_name(f"{base_name}_{idx}.xlsx")
                    if not candidate.exists():
                        out_path = candidate
                        break
            out_wb.save(out_path)
            self.log_to_preview(f"小米有品处理完成，共导出 {len(rows_out)} 条记录 → {out_path}")
            messagebox.showinfo("成功", f"小米有品开票文件已生成：\n{out_path}")
        except ImportError:
            messagebox.showerror("错误", "openpyxl 模块未安装！\n\n请先安装：pip install openpyxl")
        except Exception as e:
            self.log_to_preview(f"小米有品处理失败：{e}")
            messagebox.showerror("错误", f"小米有品处理失败：{e}")

    def _normalize_name(self, name: str) -> str:
        """规范化姓名：去掉括号及其中内容，去空格。"""
        import re
        if not name:
            return ""
        s = str(name).strip()
        # 去掉括号及括号中的内容（中英文括号）
        s = re.sub(r"[（(].*?[）)]", "", s)
        return s.strip()

    def process_vip_folder(self):
        """整理唯品会发票文件夹：解析XML并按姓名+金额匹配订单号。"""
        if not self.excel_file:
            messagebox.showwarning("提示", "请先选择唯品会表1（包含订单号、发票抬头、发票金额）")
            return
        
        from tkinter import filedialog
        folder = filedialog.askdirectory(title="选择唯品会发票文件夹（包含XML/PDF）")
        if not folder:
            return
        
        try:
            import xml.etree.ElementTree as ET
            from openpyxl import load_workbook, Workbook
            from datetime import date
            from pathlib import Path
            import os
            
            folder_path = Path(folder)
            xlsx_path = Path(self.excel_file)
            self.log_to_preview(f"使用表1：{xlsx_path}", clear=True)
            self.log_to_preview(f"扫描发票文件夹：{folder_path}")
            
            # 读取表1作为订单来源
            wb_src, ws_src = self._load_worksheet(xlsx_path)
            header_map = {}
            for col in range(1, ws_src.max_column + 1):
                v = ws_src.cell(row=1, column=col).value
                if v is not None:
                    header_map[str(v).strip()] = col
            
            def col_src(name):
                return header_map.get(name)
            
            required_cols = ["订单号", "发票抬头", "发票金额"]
            missing = [c for c in required_cols if c not in header_map]
            if missing:
                raise ValueError(f"表1缺少列：{', '.join(missing)}")
            
            # 构建 (姓名规范化, 金额) -> 订单号 列表（可能多条）
            order_index = {}
            for r in range(2, ws_src.max_row + 1):
                order_no = ws_src.cell(row=r, column=col_src("订单号")).value
                name = ws_src.cell(row=r, column=col_src("发票抬头")).value or ""
                amt = ws_src.cell(row=r, column=col_src("发票金额")).value
                if order_no is None or amt is None:
                    continue
                try:
                    amt_f = float(amt)
                except Exception:
                    continue
                key = (self._normalize_name(name), round(amt_f, 2))
                order_index.setdefault(key, []).append(str(order_no))
            
            # 扫描XML文件
            rows_out = []
            for xml_file in folder_path.rglob("*.xml"):
                try:
                    tree = ET.parse(xml_file)
                    root = tree.getroot()
                except Exception as e:
                    self.log_to_preview(f"解析XML失败：{xml_file.name} - {e}")
                    continue
                
                # XML结构假定为 <EInvoice> 根节点
                # 发票抬头
                buyer_name = root.findtext("./EInvoiceData/BuyerInformation/BuyerName", default="")
                buyer_norm = self._normalize_name(buyer_name)
                
                # 金额（含税）
                amt_text = root.findtext("./EInvoiceData/BasicInformation/TotalTax-includedAmount", default="0")
                try:
                    amt_f = float(amt_text)
                except Exception:
                    amt_f = 0.0
                amt_f = round(amt_f, 2)
                
                # 发票号码
                invoice_no = root.findtext("./TaxSupervisionInfo/InvoiceNumber", default="")
                # 发票代码：先用卖家税号
                seller_tax = root.findtext("./EInvoiceData/SellerInformation/SellerIdNum", default="")
                
                # 发票类型
                inv_type_label = root.findtext("./Header/InherentLabel/EInvoiceType/LabelName", default="")
                # 根据LabelName粗略映射到四种之一
                if "电子" in inv_type_label and "全电" not in inv_type_label:
                    invoice_type = "电子发票"
                else:
                    # 预留：遇到全电XML时再细化
                    invoice_type = inv_type_label or "电子发票"
                
                # 发票创建方式
                issue_way = ""
                for label in root.findall("./Header/UndefinedLabel/Label"):
                    ltype = label.findtext("LabelType", default="")
                    if "发票开具方式" in str(ltype):
                        issue_way = label.findtext("LabelName", default="")
                        break
                if not issue_way:
                    issue_way = "正常"
                
                # 开票内容
                item_name = root.findtext("./EInvoiceData/IssuItemInformation/ItemName", default="")
                # 去掉第一个前导星号
                if item_name.startswith("*"):
                    item_name_clean = item_name[1:]
                else:
                    item_name_clean = item_name
                content_type = "商品类别"
                content_detail = item_name_clean
                
                # 是否红冲发票（当前XML看不出红冲信息，统一否）
                is_red = "否"
                
                # 匹配订单号：按 姓名规范化 + 金额
                key = (buyer_norm, amt_f)
                order_no = ""
                remark = ""
                orders = order_index.get(key)
                if orders is None:
                    remark = "未找到匹配订单号"
                elif len(orders) == 1:
                    order_no = orders[0]
                else:
                    remark = f"同名同金额多笔，候选订单号: {','.join(orders)}"
                
                # 固定导出表中的发票类型为“全电普票”（B列）
                invoice_type = "全电普票"

                rows_out.append([
                    order_no,                    # 订单号
                    invoice_type,                # 发票类型
                    buyer_norm or buyer_name,    # 发票抬头
                    invoice_no,                  # 发票号码
                    seller_tax,                  # 发票代码（暂用卖家税号）
                    "",                         # xml链接（按要求留空）
                    "",                         # 发票下载链接
                    content_type,                # 开票内容类型
                    content_detail,              # 开票内容明细
                    is_red,                      # 是否红冲发票
                    "",                         # 配送公司
                    "",                         # 发票运单号
                    "",                         # 原发票号码
                    "",                         # 原发票代码
                    amt_f,                       # 金额
                    issue_way,                   # 发票创建方式
                    remark,                      # 备注（用于提示匹配情况）
                ])
            
            if not rows_out:
                messagebox.showwarning("提示", "未在文件夹中解析到任何有效XML发票")
                self.log_to_preview("未解析到任何发票记录，未生成文件。")
                return
            
            # 导出汇总表
            out_wb = Workbook()
            out_ws = out_wb.active
            headers = [
                "订单号",
                "发票类型",
                "发票抬头",
                "发票号码",
                "发票代码",
                "xml链接",
                "发票下载链接",
                "开票内容类型",
                "开票内容明细",
                "是否红冲发票",
                "配送公司",
                "发票运单号",
                "原发票号码",
                "原发票代码",
                "金额",
                "发票创建方式",
                "备注",
            ]
            out_ws.append(headers)
            for row in rows_out:
                out_ws.append(row)
            
            # 附带原始表1和文件列表在后续sheet
            src_sheet = out_wb.create_sheet("表1_原始")
            for src_row in ws_src.iter_rows(values_only=True):
                src_sheet.append(list(src_row))
            
            file_sheet = out_wb.create_sheet("文件列表")
            file_sheet.append(["文件名", "完整路径"])
            for xml_file in folder_path.rglob("*.xml"):
                file_sheet.append([xml_file.name, str(xml_file)])
            
            today = date.today()
            base_name = f"唯品会发票汇总_{today.year}{today.month:02d}{today.day:02d}"
            out_path = folder_path / f"{base_name}.xlsx"
            if out_path.exists():
                for idx in range(2, 100):
                    candidate = folder_path / f"{base_name}_{idx}.xlsx"
                    if not candidate.exists():
                        out_path = candidate
                        break
            out_wb.save(out_path)
            self.log_to_preview(f"唯品会发票文件夹整理完成，共 {len(rows_out)} 条记录 → {out_path}")
            messagebox.showinfo("成功", f"唯品会发票汇总已生成：\n{out_path}")
        except Exception as e:
            self.log_to_preview(f"整理唯品会发票文件夹失败：{e}")
            messagebox.showerror("错误", f"整理唯品会发票文件夹失败：{e}")
    
    def toggle_theme(self):
        """切换主题（保留数据）"""
        from theme import get_theme
        current_theme = get_theme()
        new_dark_mode = not current_theme.is_dark
        
        # 保存数据
        saved_file = self.excel_file
        
        self.root.destroy()
        import theme
        theme._global_theme = theme.RazerTheme(dark_mode=new_dark_mode)
        
        new_root = tk.Tk()
        app = InvoiceProcessorApp(new_root)
        
        # 恢复数据
        if saved_file:
            app.excel_file = saved_file
            app.file_label.config(
                text=f"✓ {Path(saved_file).name}",
                fg=app.colors['text_primary'],
                font=("Microsoft YaHei UI", 10, "bold")
            )
            app.generate_btn.config_state("normal")
            app.clear_btn.config_state("normal")
            app.update_preview()
        
        new_root.mainloop()
    
    def back_to_launcher(self):
        """返回主界面"""
        self.root.destroy()
        import sys
        sys.path.insert(0, str(Path(__file__).parent.parent.parent))
        from launcher import ToolLauncher
        new_root = tk.Tk()
        app = ToolLauncher(new_root)
        new_root.mainloop()


if __name__ == "__main__":
    root = tk.Tk()
    app = InvoiceProcessorApp(root)
    root.mainloop()

