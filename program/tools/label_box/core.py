# -*- coding: utf-8 -*-
"""
标签箱唛_螺旋桨映射版.py — v2025-09-02-merge-LB-BOX-gz-safe-logsdir + propeller-map

新增要点：
- 新增 PROPELLER_MAP：当配货表中 E 列为“螺旋桨”时，优先按 A 列商品编号映射到固定的螺旋桨 .pld 模板。
- 新增统一读取函数 read_id_sku_e_from_sheet：同时返回 A 列(商品编号)、E 列(品名/类型) 与用于匹配的 ID 字段。
- 在生成标签的匹配循环中加入“螺旋桨特判”，命中映射则直接强制使用对应 .pld，跳过通用候选逻辑。

其余逻辑保持与原版一致：
- 统一日志目录 <脚本同级>/日志
- 标签拷贝、.pld 日期批改、箱唛批处理等
"""

import re, sys, traceback, shutil, os, datetime
from datetime import datetime as dt
from pathlib import Path

# ---------- GUI ----------
try:
    import tkinter as tk
    from tkinter import filedialog, messagebox
    TK_OK = True
except Exception:
    TK_OK = False

try:
    from openpyxl import load_workbook
except ImportError:
    if TK_OK:
        tk.Tk().withdraw()
        messagebox.showerror("缺少依赖", "未找到 openpyxl，请先安装：\npython -m pip install openpyxl")
    else:
        print("缺少依赖：openpyxl。请先安装：python -m pip install openpyxl")
    sys.exit(1)

VERSION = "v2025-09-02-merge-LB-BOX-gz-safe-logsdir+propeller-map+use-拆1+patch-Eonly"

# ===================== 通用与标签侧工具 =====================
def strip_norm(s: str) -> str:
    return (s or "").strip()

def to_halfwidth(s: str) -> str:
    res = []
    for ch in str(s):
        code = ord(ch)
        if code == 0x3000:
            res.append(" ")
        elif 0xFF01 <= code <= 0xFF5E:
            res.append(chr(code - 0xFEE0))
        else:
            res.append(ch)
    return "".join(res)

def looks_like_mmdd(s4: str) -> bool:
    if not re.fullmatch(r"\d{4}", s4): return False
    mm = int(s4[:2]); dd = int(s4[2:])
    return 1 <= mm <= 12 and 1 <= dd <= 31

EXPECTED_SHEETS = ["外仓库配货表", "梨配货表", "兽仓库配货表", "兽无人机拆1", "兽无人机拆2", "兽无人机仓库配货表总"]
SHEET_ALIASES = {
    "外仓库配货表": ["外星人配货表", "外仓库配货表"],
    "梨配货表":     ["梨配货表"],
    "兽仓库配货表": ["兽配货表", "兽仓库配货表"],
    "兽无人机拆1": ["兽无人机拆1"],
    "兽无人机拆2": ["兽无人机拆2"],
    "兽无人机仓库配货表总": ["兽无人机仓库配货表总"],
}


# 对这些表：仅用 E 列匹配（禁用 A 列/数字前缀兜底，保留“螺旋桨”特判）
FORCE_USE_E_ONLY_SHEETS = {"外仓库配货表"}

def choose_workbook(initial_dir: Path) -> Path:
    if not TK_OK:
        raise RuntimeError("当前环境无图形界面，无法选择 Excel。")
    root = tk.Tk(); root.withdraw()
    file_path = filedialog.askopenfilename(
        title="请选择含配货/箱唛信息的 Excel 文件（同一工作簿）",
        initialdir=str(initial_dir),
        filetypes=[("Excel 文件", "*.xlsx *.xlsm *.xltx *.xltm"), ("所有文件", "*.*")]
    )
    root.update(); root.destroy()
    if not file_path:
        raise RuntimeError("未选择 Excel 文件。")
    return Path(file_path)

def resolve_sheet_names(wb) -> dict:
    mapping = {}
    names = wb.sheetnames
    def norm(s): return strip_norm(s)
    for exp in EXPECTED_SHEETS:
        hit = None
        for alias in SHEET_ALIASES.get(exp, [exp]):
            for real in names:
                if norm(real) == alias:
                    hit = real; break
            if hit: break
        mapping[exp] = hit
    return mapping

# --- 统一读取：返回 (id_from_idcol, sku_from_A, e_value_from_E) 三元组 ---
# 说明：
# - id_col: 用于现有匹配逻辑的“编号/关键字”列（原脚本中默认为 E 列=5）
# - sku_col: A 列商品编号（本次用于“螺旋桨特判”的映射键）
# - e_col: E 列文本（用来判断是否为“螺旋桨”）

def read_id_sku_e_from_sheet(wb, real_sheet_name: str, id_col: int = 5, sku_col: int = 1, e_col: int = 5, start_row: int = 2):
    ws = wb[real_sheet_name]
    rows, seen = [], set()
    for row_num, r in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
        idv = r[id_col - 1].value
        skuv = r[sku_col - 1].value
        e_val = r[e_col - 1].value
        if idv is None and skuv is None and e_val is None:
            continue
        idstr = str(idv).strip() if idv is not None else ""
        sku = str(skuv).strip() if skuv is not None else ""
        e_txt = str(e_val).strip() if e_val is not None else ""
        
        # 过滤：只保留 A 列（SKU）是数字的行
        # 这样可以排除统计行、合计行等非数据行
        sku_is_number = sku and sku.isdigit()
        
        key = (idstr, sku, e_txt)
        # 只有当 SKU 是数字时，才添加到结果中
        if sku_is_number and key not in seen:
            rows.append((idstr, sku, e_txt, row_num)); seen.add(key)
    return rows

def read_b1(wb, real_sheet_name: str) -> str:
    try:
        ws = wb[real_sheet_name]
        v = ws["B1"].value
        if v is None: return ""
        return str(v).strip()
    except Exception:
        return ""

def extract_mmdd_from_text_window(text: str):
    digits = "".join(ch for ch in str(text) if ch.isdigit())
    for i in range(0, max(0, len(digits) - 3)):
        sub = digits[i:i+4]
        if looks_like_mmdd(sub):
            return sub
    return None

def extract_mmdd_from_sn(sn: str):
    s = to_halfwidth(sn).strip()
    if len(s) >= 7:
        sub = s[-7:-3]
        if looks_like_mmdd(sub):
            return sub
    return extract_mmdd_from_text_window(s)

DATE_SRC_NOTE = ""

def extract_date_tag_from_wb(wb, sheet_name_map: dict) -> str:
    """遍历核心表 D 列(SN) 提取 MMDD；找不到回退当天。"""
    global DATE_SRC_NOTE
    for exp in EXPECTED_SHEETS:
        real = sheet_name_map.get(exp)
        if not real: continue
        ws = wb[real]
        for idx, row in enumerate(ws.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True), start=2):
            cell = row[0]
            if cell is None: continue
            mmdd = extract_mmdd_from_sn(str(cell))
            if mmdd:
                DATE_SRC_NOTE = f"{real} 第{idx}行 D 列：{cell} → {mmdd}"
                return mmdd
    mmdd_fallback = datetime.date.today().strftime("%m%d")
    DATE_SRC_NOTE = f"未在 D 列 SN 中找到，回退当天：{mmdd_fallback}"
    return mmdd_fallback

def _has_3c_token(text: str) -> bool:
    s = to_halfwidth(text or "").lower()
    return ("3c" in s) or ("3 c" in s)

def _has_toy_token(text: str) -> bool:
    s = to_halfwidth(text or "").lower()
    return ("玩具" in s) or ("toy" in s)

def decide_label_type_by_b1(b1_values: dict, sheet_name_map: dict) -> str:
    """3C优先；空值不投票；无明确信号默认 3C。"""
    votes_3c = votes_toy = 0
    for exp_sheet, _real in sheet_name_map.items():
        b1 = (b1_values.get(exp_sheet) or "").strip()
        if not b1: continue
        if _has_3c_token(b1): votes_3c += 1
        if _has_toy_token(b1): votes_toy += 1
    if votes_3c and not votes_toy: return "3C"
    if votes_toy and not votes_3c: return "玩具"
    if votes_3c and votes_toy: return "3C"
    return "3C"

def resolve_label_template_dir(root_dir: Path, dirname: str):
    """查找 标签模板 目录（多路径/同级自动搜寻）"""
    tried = []; seen=set()
    def add(p: Path):
        sp=str(p); 
        if sp not in seen: tried.append(sp); seen.add(sp)
        return p
    cands = [
        add(root_dir / dirname),
        add(root_dir / "标签模板" / dirname),
        add(root_dir / "标签模板文件夹" / dirname),
    ]
    try:
        for d in root_dir.iterdir():
            if d.is_dir():
                cands.append(add(d / dirname))
    except Exception:
        pass
    for p in cands:
        if p and p.exists() and p.is_dir():
            return p, tried
    return None, tried

def sheet_prefix_of(sheet: str) -> str:
    return {
        "外仓库配货表": "外星人",
        "梨配货表": "三只梨",
        "兽仓库配货表": "兽",
        "兽无人机拆1": "兽无人机",
    }.get(sheet, "")

def build_variants(raw_id: str):
    def normalize_basic(raw: str):
        s0=str(raw); s=s0
        if "售止" in s: s=s.replace("售止","")
        s=to_halfwidth(s).strip().replace(" ","")
        for k,v in {"螺旋奖":"螺旋桨","螺施桨":"螺旋桨","螺桨":"螺旋桨"}.items():
            s=s.replace(k,v)
        return s
    def strip_tail_plus_num(s: str): return re.sub(r"\+\d+$","",s)
    def remove_all_plus(s: str): return s.replace("+","")
    vs, seen=[], set()
    def add(v):
        if v not in seen: vs.append(v); seen.add(v)
    base=normalize_basic(raw_id); add(raw_id); add(base)
    add(strip_tail_plus_num(base)); add(remove_all_plus(base)); add(remove_all_plus(strip_tail_plus_num(base)))
    return vs

def candidate_filenames(sheet: str, id_variants: list, forced_names: list = None) -> list:
    prefix = sheet_prefix_of(sheet)
    cands,seen=[],set()
    def add(name):
        if name not in seen: cands.append(name); seen.add(name)
    
    # 优先处理强制指定的文件名（螺旋桨映射）
    if forced_names:
        for nm in forced_names:
            # nm 可传入不带后缀或已带 .pld，这里统一两种情况
            if nm.endswith('.pld'):
                # 直接添加，不做任何修改（保留原始文件名，包括+号）
                add(nm)
                # 只有当前缀存在且不在名称开头时，才添加带前缀的版本
                if prefix and not nm.startswith(prefix):
                    add(f"{prefix}{nm}")
            else:
                add(f"{nm}.pld")
                if prefix and not f"{nm}.pld".startswith(prefix):
                    add(f"{prefix}{nm}.pld")
    
    # 然后处理ID变体（通用匹配）
    for the_id in id_variants:
        if the_id:
            add(f"{the_id}.pld")
            if prefix: add(f"{prefix}{the_id}.pld")
    return cands

def build_pld_index(base_dir: Path):
    idx = {}; entries=[]
    for p in base_dir.rglob("*.pld"):
        if p.is_file():
            name_l = p.name.lower()
            if name_l not in idx:
                idx[name_l]=p
                entries.append((p.stem.lower(), name_l, p))
    idx.pop("拷贝结果日志.txt".lower(), None)
    entries=[e for e in entries if e[1] != "拷贝结果日志.txt".lower()]
    return idx, entries

def find_matching_templates(index_lower: dict, cand_names: list):
    matches,seen=[],set()
    for name in cand_names:
        p = index_lower.get(name.lower())
        if p and p not in seen:
            matches.append(p); seen.add(p)
    return matches

def numeric_prefix(s: str) -> str:
    m = re.match(r"(\d+)", s)
    return m.group(1) if m else ""

def fallback_by_number_prefix(entries, number_prefix: str, sheet: str):
    prefix_sheet = sheet_prefix_of(sheet).lower()
    hits,seen=[],set()
    for stem_l, name_l, p in entries:
        ok=False
        if stem_l.startswith(number_prefix): ok=True
        elif prefix_sheet and stem_l.startswith(prefix_sheet) and stem_l[len(prefix_sheet):].startswith(number_prefix):
            ok=True
        if ok and p not in seen:
            hits.append(p); seen.add(p)
    return hits

# === .pld 日期批改（标签侧） ===
def is_ascii_printable(b: int) -> bool: return 32 <= b <= 126

def find_ascii_spans(data: bytes, min_len: int = 6):
    spans=[]; n=len(data); i=0
    while i<n:
        if is_ascii_printable(data[i]):
            j=i+1
            while j<n and is_ascii_printable(data[j]): j+=1
            if j-i>=min_len: spans.append((i,j-1))
            i=j
        else:
            i+=1
    return spans

KEYWORDS_NEAR_DATE = ("----", "SN", "序列号", "SN序列号", "日期")

def looks_mmdd_bytes(mm4b: bytes) -> bool:
    try:
        s=mm4b.decode("ascii",errors="ignore")
        return looks_like_mmdd(s)
    except Exception:
        return False

def replace_independent_mmdd_in_block(buf: bytearray, start: int, end: int, new_mmdd: bytes) -> int:
    changed=0
    block=buf[start:end+1]
    text=block.decode('latin1', errors='ignore')
    for m in re.finditer(r"(?<!\d)(\d{4})(?!\d)", text):
        mm=m.group(1)
        if looks_like_mmdd(mm):
            abs_pos = start + m.start(1)
            left=max(0,m.start(1)-80); right=min(len(text),m.end(1)+80)
            ctx=text[left:right]
            if any(k in ctx for k in KEYWORDS_NEAR_DATE):
                buf[abs_pos:abs_pos+4]=new_mmdd; changed+=1
    return changed

def replace_corner_mmdd_in_block(buf: bytearray, start: int, end: int, new_mmdd: bytes) -> int:
    changed=0
    block=buf[start:end+1]
    text=block.decode('latin1', errors='ignore')
    for m in re.finditer(r"(?<!\d)(\d{4})(?!\d)", text):
        mm=m.group(1)
        if looks_like_mmdd(mm):
            abs_pos = start + m.start(1)
            buf[abs_pos:abs_pos+4]=new_mmdd; changed+=1
    return changed

def replace_sn_mmdd_in_block(buf: bytearray, start: int, end: int, new_mmdd: bytes) -> int:
    changed=0
    block=buf[start:end+1]
    text=block.decode('latin1', errors='ignore')
    for m in re.finditer(r'([A-Za-z]{1,10})(\d{10,})', text):
        digits=m.group(2)
        if len(digits) >= 7:
            mmdd_pos_start = len(digits) - 7
            abs_mmdd_start = start + m.start(2) + mmdd_pos_start
            buf[abs_mmdd_start:abs_mmdd_start+4]=new_mmdd; changed+=1
    return changed

def replace_any_standalone_mmdd_bytes(buf: bytearray, new_mmdd: bytes) -> int:
    changed=0
    for m in re.finditer(rb'(?<![0-9])([0-9]{4})(?![0-9])', buf):
        start=m.start(1)
        if looks_mmdd_bytes(buf[start:start+4]):
            buf[start:start+4]=new_mmdd; changed+=1
    return changed

def process_pld_file(path: Path, new_mmdd: str, dry_run: bool, make_backup: bool) -> dict:
    data = path.read_bytes()
    buf = bytearray(data)
    new_b = new_mmdd.encode('ascii')

    spans = find_ascii_spans(data, min_len=6)
    sn_changes = text_changes = corner_changes = 0
    for (s, e) in spans:
        sn_changes     += replace_sn_mmdd_in_block(buf, s, e, new_b)
        text_changes   += replace_independent_mmdd_in_block(buf, s, e, new_b)
        corner_changes += replace_corner_mmdd_in_block(buf, s, e, new_b)

    bytes_changes = replace_any_standalone_mmdd_bytes(buf, new_b)

    changed = (sn_changes + text_changes + corner_changes + bytes_changes) > 0
    wrote = False
    bak_path = None
    if changed and not dry_run:
        if make_backup:
            bak_path = path.with_suffix(path.suffix + ".bak")
            if not bak_path.exists():
                bak_path.write_bytes(data)
        path.write_bytes(buf)
        wrote = True

    return {
        "file": str(path),
        "sn_changes": sn_changes,
        "text_changes": text_changes,
        "corner_changes": corner_changes,
        "bytes_changes": bytes_changes,
        "changed": changed,
        "wrote": wrote,
        "bak": str(bak_path) if bak_path else ""
    }

def run_patch_step(base_dir: Path, mmdd: str, ext: str = ".pld", dry: bool = False, make_backup: bool = False, report_dir: Path = None):
    targets = sorted(Path(base_dir).rglob(f"*{ext}"))
    rows = []
    tot_sn = tot_text = tot_corner = tot_bytes = tot_files = tot_wrote = 0
    for p in targets:
        try:
            res = process_pld_file(p, mmdd, dry, make_backup)
            rows.append(res)
            tot_sn     += res["sn_changes"]
            tot_text   += res["text_changes"]
            tot_corner += res["corner_changes"]
            tot_bytes  += res["bytes_changes"]
            tot_files  += 1
            tot_wrote  += 1 if res["wrote"] else 0
        except PermissionError:
            rows.append({"file": str(p), "error": "无权访问（可能被占用）"})
        except Exception as e:
            rows.append({"file": str(p), "error": str(e)})
    summary = (
        f"目录：{base_dir}\n"
        f"目标MMDD：{mmdd}\n"
        f"模式：{'预览(不写回)' if dry else '写回'}；备份：{'开' if make_backup and not dry else '关或预览'}\n"
        f"共处理文件：{tot_files}\n"
        f"SN改：{tot_sn}  |  关键字/整行改：{tot_text}  |  独立4位改：{tot_corner}  |  字节兜底改：{tot_bytes}\n"
        f"实际写入文件数：{tot_wrote}\n"
    )
    # 构建报告内容
    report_lines = []
    report_lines.append(f"[批处理时间] {dt.now().strftime('%Y-%m-%d %H:%M:%S')}")
    report_lines.append(summary)
    for r in rows:
        if "error" in r:
            report_lines.append(f"[错误] {r['file']} -> {r['error']}")
        else:
            report_lines.append(f"[OK] {r['file']} | SN:{r['sn_changes']}  文本:{r['text_changes']}  角标:{r['corner_changes']}  字节:{r['bytes_changes']}  写入:{r['wrote']}  备份:{r['bak']}")
    
    # 组合报告文本
    report_text = "\n".join(report_lines)
    
    # 如果指定了report_dir，保存文件
    if report_dir is not None:
        report_dir.mkdir(parents=True, exist_ok=True)
        report_path = report_dir / "patch_pld_report.txt"
        with report_path.open("w", encoding="utf-8") as f:
            f.write(report_text)
        return summary, report_path
    else:
        # 不保存文件，只返回报告文本
        return summary, report_text

# ===================== 螺旋桨映射（增强版） =====================
# 导入螺旋桨配置
try:
    from .propeller_config import (
        STATIC_PROPELLER_MAP as PROPELLER_MAP,
        PROPELLER_KEYWORDS,
        SHOP_PREFIXES,
        SHEET_TO_SHOP_MAP
    )
except ImportError:
    # 向后兼容：如果配置文件不存在，使用默认映射
    PROPELLER_MAP = {
        "100181107889": "外星人螺旋桨.pld",
        "100235985474": "三只梨螺旋桨.pld",
        "100264779838": "兽螺旋桨.pld",
        "100144781118": "901螺旋桨.pld",
        "100131174559": "909螺旋桨.pld",
    }
    PROPELLER_KEYWORDS = ["螺旋桨", "propeller", "螺桨", "螺旋奖", "螺施桨"]
    SHOP_PREFIXES = ["外星人", "三只梨", "兽", "901", "909"]
    SHEET_TO_SHOP_MAP = {
        "外仓库配货表": "外星人",
        "梨配货表": "三只梨", 
        "兽仓库配货表": "兽",
        "兽无人机拆1": "兽",
        "兽无人机拆2": "兽",
    }

def build_dynamic_propeller_map(pld_index, pld_entries):
    """
    动态构建螺旋桨映射表，自动发现所有螺旋桨相关的PLD文件
    
    参数:
        pld_index: PLD文件索引字典 {filename_lower: Path}
        pld_entries: PLD文件条目列表 [(stem_lower, name_lower, Path)]
    
    返回:
        动态螺旋桨映射字典 {商品编号: pld文件名}
    """
    import json
    import re
    from pathlib import Path
    
    dynamic_map = {}
    
    # 1. 首先尝试从 propeller_mappings.json 中读取映射表
    try:
        mappings_file = Path(__file__).parent / "propeller_mappings.json"
        if mappings_file.exists():
            with open(mappings_file, 'r', encoding='utf-8') as f:
                json_mappings = json.load(f)
                if isinstance(json_mappings, dict):
                    # 直接添加 JSON 中的所有映射
                    for sku, pld_name in json_mappings.items():
                        dynamic_map[sku] = pld_name
    except Exception as e:
        pass  # 如果读取失败，继续使用其他方法
    
    # 2. 遍历所有PLD文件，查找螺旋桨相关文件
    for stem_lower, name_lower, pld_path in pld_entries:
        filename = pld_path.name
        filename_lower = filename.lower()
        
        # 检查文件名是否包含螺旋桨关键词
        is_propeller_file = any(keyword in filename for keyword in PROPELLER_KEYWORDS)
        
        if is_propeller_file:
            # 模式1: 直接的数字编号 (如: 100181107889螺旋桨.pld)
            number_match = re.search(r'(\d{10,})', filename)
            if number_match:
                product_code = number_match.group(1)
                # 只有当 JSON 中没有这个映射时，才使用从文件名提取的
                if product_code not in dynamic_map:
                    dynamic_map[product_code] = filename
                continue
            
            # 模式2: 店铺前缀+螺旋桨 (如: 外星人螺旋桨.pld, 三只梨螺旋桨.pld)
            for prefix in SHOP_PREFIXES:
                if prefix in filename:
                    shop_key = f"店铺_{prefix}"
                    # 只有当 JSON 中没有这个映射时，才使用从文件名提取的
                    if shop_key not in dynamic_map:
                        dynamic_map[shop_key] = filename
                    break
    
    return dynamic_map

def find_propeller_template(sku_str, raw_id_str, exp_sheet, pld_index, pld_entries):
    """
    增强的螺旋桨模板查找函数
    
    参数:
        sku_str: A列商品编号
        raw_id_str: E列ID
        exp_sheet: 工作表名称
        pld_index: PLD文件索引
        pld_entries: PLD文件条目列表
    
    返回:
        匹配的PLD文件名，如果没找到返回None
    """
    # 1. 首先尝试静态映射表（保持向后兼容）
    forced_name = PROPELLER_MAP.get(sku_str) or PROPELLER_MAP.get(raw_id_str)
    if forced_name:
        return forced_name
    
    # 2. 构建动态映射表
    dynamic_map = build_dynamic_propeller_map(pld_index, pld_entries)
    
    # 3. 尝试商品编号直接匹配
    if sku_str and sku_str in dynamic_map:
        return dynamic_map[sku_str]
    
    if raw_id_str and raw_id_str in dynamic_map:
        return dynamic_map[raw_id_str]
    
    # 4. 尝试根据工作表名称匹配店铺螺旋桨
    shop_name = SHEET_TO_SHOP_MAP.get(exp_sheet)
    if shop_name:
        shop_key = f"店铺_{shop_name}"
        if shop_key in dynamic_map:
            return dynamic_map[shop_key]
    
    # 5. 模糊匹配：查找包含商品编号的螺旋桨文件
    if sku_str and len(sku_str) >= 6:  # 商品编号通常比较长
        for filename in pld_index.keys():
            if "螺旋桨" in filename and sku_str in filename:
                return pld_index[filename].name
    
    # 6. 最后尝试通用螺旋桨文件（按店铺优先级）
    propeller_files = []
    for stem_lower, name_lower, pld_path in pld_entries:
        if "螺旋桨" in pld_path.name:
            propeller_files.append(pld_path.name)
    
    if propeller_files:
        # 根据工作表优先选择对应店铺的螺旋桨文件
        if shop_name:
            for filename in propeller_files:
                if shop_name in filename:
                    return filename
        
        # 如果没有找到对应店铺的，返回第一个螺旋桨文件
        return propeller_files[0]
    
    return None

# ===================== 箱唛侧工具 =====================
FW = "\u3000"  # 全角空格 U+3000
CITY_KEYS = ["北京","上海","广州","成都","武汉","沈阳","西安","德州"]
CITY_META = {
    "北京": {"province":"",   "prov_disp":"",        "city_only":"北京市", "combined":"北京市"},
    "上海": {"province":"",   "prov_disp":"",        "city_only":"上海市", "combined":"上海市"},
    "广州": {"province":"广东","prov_disp":"广东"+FW, "city_only":"广州市", "combined":"广东"+FW+"广州市"},
    "成都": {"province":"四川","prov_disp":"四川"+FW, "city_only":"成都市", "combined":"四川"+FW+"成都市"},
    "武汉": {"province":"湖北","prov_disp":"湖北"+FW, "city_only":"武汉市", "combined":"湖北"+FW+"武汉市"},
    "沈阳": {"province":"辽宁","prov_disp":"辽宁"+FW, "city_only":"沈阳市", "combined":"辽宁"+FW+"沈阳市"},
    "西安": {"province":"陕西","prov_disp":"陕西"+FW, "city_only":"西安市", "combined":"陕西"+FW+"西安市"},
    "德州": {"province":"山东","prov_disp":"山东"+FW, "city_only":"德州市", "combined":"山东"+FW+"德州市"},
}
CITY_ALTS_COMBINED = {
    "北京": ["北京市"], "上海": ["上海市"],
    "广州": ["广东"+FW+"广州市", "广东 广州市"],
    "成都": ["四川"+FW+"成都市", "四川 成都市"],
    "武汉": ["湖北"+FW+"武汉市", "湖北 武汉市"],
    "沈阳": ["辽宁"+FW+"沈阳市", "辽宁 沯阳市".replace("沯","阳")],
    "西安": ["陕西"+FW+"西安市", "陕西 西安市"],
    "德州": ["山东"+FW+"德州市", "山东 德州市"],
}
CITY_ALTS_CITYONLY = {
    "北京": ["北京市"], "上海": ["上海市"],
    "广州": ["广州市"], "成都": ["成都市"], "武汉": ["武汉市"],
    "沈阳": ["沈阳市"], "西安": ["西安市"], "德州": ["德州市"],
}
CITY_ALTS_PROVONLY = {
    "广州": ["广东"+FW, "广东 "],
    "成都": ["四川"+FW, "四川 "],
    "武汉": ["湖北"+FW, "湖北 "],
    "沈阳": ["辽宁"+FW, "辽宁 "],
    "西安": ["陕西"+FW, "陕西 "],
    "德州": ["山东"+FW, "山东 "],
}
STORE_KEYS = ["三只梨", "兽无人机", "外星人", "兽"]

def find_box_sheet(wb):
    visible = [ws for ws in wb.worksheets if ws.sheet_state == "visible"]
    for ws in visible:
        if ws.title.strip() == "箱唛":
            return ws
    for ws in visible:
        if "箱唛" in ws.title:
            return ws
    return visible[0] if visible else None

def iter_visible_rows(ws):
    """只迭代未隐藏的行号"""
    for row_idx in range(1, ws.max_row + 1):
        # 注意：openpyxl 中可见行 often 是 None 或 False；只有 True 才是隐藏
        if ws.row_dimensions.get(row_idx) and ws.row_dimensions[row_idx].hidden is True:
            continue
        yield row_idx

def decide_store_subfolder(ws) -> str or None:
    hdr = ""
    for r in range(1, 4):
        row_vals = [str(ws.cell(row=r, column=c).value or "") for c in range(1, 5+1)]
        hdr += " ".join(row_vals) + " "
    if "店箱唛" in hdr:   # 两店/四店等：不建子目录
        return None
    for key in STORE_KEYS:
        if key in hdr:
            return key
    return None

def parse_entries(ws):
    entries, cur = [], None

    def commit():
        nonlocal cur
        if cur and cur.get("city") and cur.get("supplier") and cur.get("po") and cur.get("depot"):
            entries.append(cur)
        cur = None

    for r in iter_visible_rows(ws):
        row_vals = [str(ws.cell(r, c).value or "").strip() for c in range(1, 10)]
        line = " ".join(v for v in row_vals if v)

        if "目的地" in line:
            commit()
            city = next((ck for ck in ["北京","上海","广州","成都","武汉","沈阳","西安","德州"] if ck in line), None)
            cur = {"city": city, "po": None, "supplier": None, "depot": None, "no": None}
            m = re.search(r"序号\D*(\d{1,3})", line)
            if m: cur["no"] = m.group(1)
            continue

        if cur is None: 
            continue

        if ("供应商简码" in line) or ("供应商代码" in line) or ("商家名称" in line):
            cand = [t for t in row_vals if t][-1]
            cur["supplier"] = re.sub(r"[^A-Za-z0-9_\-\.]", "", cand)
            continue

        if ("采购单号" in line) or ("PO" in line.upper()):
            m = re.search(r"(\d{6,})", line) or next((re.match(r"(\d{6,})", v) for v in row_vals if re.match(r"(\d{6,})", v)), None)
            if m: cur["po"] = m.group(1)
            continue

        if ("目的库房" in line) or ("目的仓" in line):
            vals = [v for v in row_vals if v]
            if vals:
                depot = vals[-1].replace("；；", "；").strip()
                if depot != "#N/A":              # ✅ 排除无效库房
                    cur["depot"] = depot
            continue

    commit()
    return entries

    max_row = ws.max_row
    for r in range(1, max_row + 1):
        row_vals = [str(ws.cell(row=r, column=c).value or "").strip() for c in range(1, 10)]
        line = " ".join(v for v in row_vals if v)

        if "目的地" in line:
            commit()
            city = None
            for ck in CITY_KEYS:
                if ck in line:
                    city = ck; break
            cur = {"city": city, "po": None, "supplier": None, "depot": None, "no": None}
            if "序号" in line:
                m = re.search(r"序号\D*(\d{1,3})", line)
                if m: cur["no"] = m.group(1)
            continue

        if cur is None:
            continue

        if ("供应商简码" in line) or ("供应商代码" in line) or ("商家名称" in line):
            cand = [t for t in row_vals if t][-1] if any(row_vals) else ""
            last = to_halfwidth(cand).strip()
            last = re.sub(r"[^A-Za-z0-9_\-\.]", "", last)
            if last:
                cur["supplier"] = last
            continue

        if ("采购单号" in line) or ("PO" in line.upper()):
            m = re.search(r"(\d{6,})", line)
            if not m:
                for v in row_vals:
                    mv = re.match(r"(\d{6,})", v)
                    if mv: m = mv; break
            if m:
                cur["po"] = m.group(1)
            continue

        if ("目的库房" in line) or ("目的仓" in line):
            vals = [v for v in row_vals if v]
            if vals:
                cand = [str(v).replace("；；", "；").strip() for v in vals if str(v).strip() and str(v).strip() != "#N/A"]
        if cand:
            pick = next((v for v in cand if v.endswith("-CHN")), None)
            if not pick:
                pick = cand[0]
            if any(len(v) > len(pick) for v in cand):
                pick = max(cand, key=len)
            cur["depot"] = pick
            continue

        if ("序号" in line) and not cur.get("no"):
            m = re.search(r"\b(\d{1,3})\b", line)
            if m: cur["no"] = m.group(1)

    commit()
    return entries

def encode_gbk_fit(text: str, target_len: int) -> bytes:
    raw = text.encode("gbk", errors="ignore")
    if len(raw) == target_len: return raw
    if len(raw) < target_len:  return raw + b" " * (target_len - len(raw))
    cut = target_len
    while cut > 0:
        try:
            raw[:cut].decode("gbk"); break
        except UnicodeDecodeError:
            cut -= 1
    out = raw[:cut]
    if len(out) < target_len:
        out += b" " * (target_len - len(out))
    return out

def resolve_box_template_dir(root_dir: Path, kind_dirname: str):
    tried = []
    def add(p: Path): tried.append(str(p)); return p
    cand = [
        add(root_dir / "箱唛模板" / kind_dirname),
        add(root_dir / kind_dirname),
    ]
    try:
        for d in root_dir.iterdir():
            if d.is_dir():
                cand.append(add(d / "箱唛模板" / kind_dirname))
    except Exception:
        pass
    for p in cand:
        if p.exists() and p.is_dir():
            return p, tried
    return None, tried

CITY_FILE = {
    "北京": "北京箱唛.pld", "上海": "上海箱唛.pld", "广州": "广州箱唛.pld", "成都": "成都箱唛.pld",
    "武汉": "武汉箱唛.pld", "沈阳": "沈阳箱唛.pld", "西安": "西安箱唛.pld", "德州": "德州箱唛.pld",
}

def find_city_template(base_dir: Path, city: str):
    target = CITY_FILE.get(city)
    if target:
        p = base_dir / target
        if p.exists(): return p
    for q in base_dir.glob("*"):
        if not q.is_file(): continue
        name = q.name
        if (city in name) and ("箱唛" in name) and q.suffix.lower()==".pld":
            return q
        if (city in name) and ("箱唛" in name) and (".pld" in name or ",pld" in name):
            return q
    return None

# 下面箱唛 patch 逻辑与原版一致（略去注释）

def find_all_digits_regions(buf: bytes, min_len=6, max_len=20):
    return [m for m in re.finditer(rb"(?<!\d)(\d{%d,%d})(?!\d)" % (min_len, max_len), buf)]

def replace_region_bytes(buf: bytearray, start: int, old_len: int, new_bytes: bytes):
    if len(new_bytes) == old_len:
        buf[start:start+old_len] = new_bytes
    elif len(new_bytes) < old_len:
        buf[start:start+old_len] = new_bytes + b" " * (old_len - len(new_bytes))
    else:
        buf[start:start+old_len] = new_bytes[:old_len]

def search_label_positions(buf: bytes, text: str):
    hits=[]
    for enc in ("gbk","utf-16le","utf-8"):
        try:
            patt=text.encode(enc,errors="ignore")
            for m in re.finditer(re.escape(patt), buf):
                hits.append(m.start())
        except Exception:
            pass
    return sorted(set(hits))

def find_next_digits_after(buf: bytes, pos: int, max_seek=3000, min_len=1, max_len=20, find_all=False):
    region = buf[pos:pos+max_seek]
    if not find_all:
        m = re.search(rb"(?<!\d)(\d{%d,%d})(?!\d)" % (min_len, max_len), region)
        if m:
            return [(pos + m.start(1), m.group(1))]
        return []
    else:
        return [(pos + m.start(1), m.group(1)) for m in re.finditer(rb"(?<!\d)(\d{%d,%d})(?!\d)" % (min_len, max_len), region)]

def find_ascii_word_after(buf: bytes, pos: int, max_seek=3000):
    region = buf[pos:pos+max_seek]
    m = re.search(rb"([A-Za-z0-9_\-\.]{1,64})", region)
    if m:
        return pos + m.start(1), m.group(1)
    return None, None

def _gbk_find_all(buf: bytes, text: str):
    try:
        patt = text.encode("gbk", errors="ignore")
    except Exception:
        return []
    return [ (m.start(), m.end()) for m in re.finditer(re.escape(patt), buf) ]

def find_city_display_windows(buf: bytes, city: str):
    prov_hits=[]; 
    for t in CITY_ALTS_PROVONLY.get(city, []):
        prov_hits += _gbk_find_all(buf, t)
    city_hits=[]
    for t in CITY_ALTS_CITYONLY.get(city, []):
        city_hits += _gbk_find_all(buf, t)
    comb_hits=[]
    for t in CITY_ALTS_COMBINED.get(city, []):
        comb_hits += _gbk_find_all(buf, t)

    last = lambda hits: hits[-1] if hits else None
    prov_win = last(sorted(prov_hits, key=lambda x: x[0]))
    city_win = last(sorted(city_hits, key=lambda x: x[0]))
    comb_win = last(sorted(comb_hits, key=lambda x: x[0]))

    if prov_win and city_win:  return 'split',    {'prov': prov_win, 'city': city_win}
    if comb_win:               return 'combined', {'combined': comb_win}
    if city_win:               return 'city_only',{'city': city_win}
    return None, {}

def replace_bytes_window(buf: bytearray, win, new_text: str, tag: str, changed: list):
    s,e = win
    b = encode_gbk_fit(new_text, e-s)
    buf[s:e] = b
    changed.append(f"{tag}@{s}-{e}")

def replace_star_number_all(buf: bytearray, new_num: str):
    pattern = re.compile(rb"\*([0-9]{6,20})\*")
    positions=[]; cnt=0
    new_b = new_num.encode("ascii", errors="ignore")
    for m in list(pattern.finditer(buf)):
        a=m.start(1); b=m.end(1); old_len=b-a
        if len(new_b) == old_len: buf[a:b] = new_b
        elif len(new_b) < old_len: buf[a:b] = new_b + b" "*(old_len-len(new_b))
        else: buf[a:b] = new_b[:old_len]
        positions.append(a); cnt+=1
    return cnt, positions

def find_depot_region(buf: bytes, city_core: str, prefer_suffix=b"-CHN", window=1024):
    try:
        city_bytes = city_core.encode("gbk", errors="ignore")
    except Exception:
        return None, None
    last_pos = -1
    for m in re.finditer(re.escape(city_bytes), buf):
        last_pos = m.start()
    if last_pos < 0:
        return None, None
    tail = buf[last_pos:last_pos+window]
    m2 = re.search(re.escape(prefer_suffix), tail)
    if not m2:
        end_idx = last_pos + min(window, 120)
        return last_pos, end_idx
    end_idx = last_pos + m2.end()
    return last_pos, end_idx

def patch_city_display(buf: bytearray, city: str, changed: list, warns: list, debug: list):
    mode, wins = find_city_display_windows(buf, city)
    meta = CITY_META[city]
    if mode == 'split':
        replace_bytes_window(buf, wins['prov'], meta['prov_disp'], "目的城市(省段)", changed)
        replace_bytes_window(buf, wins['city'], meta['city_only'], "目的城市(市段)", changed)
        debug.append(f"city-display mode=split prov@{wins['prov']} city@{wins['city']}")
        return
    if mode == 'combined':
        replace_bytes_window(buf, wins['combined'], meta['combined'], "目的城市(合并段)", changed)
        debug.append(f"city-display mode=combined win@{wins['combined']}")
        return
    if mode == 'city_only':
        replace_bytes_window(buf, wins['city'], meta['city_only'], "目的城市(仅市段)", changed)
        warns.append("模板为仅市名段：已只写市名；如需省份请换含省段模板。")
        debug.append(f"city-display mode=city_only win@{wins['city']}")
        return
    warns.append("未定位到目的城市显示（模板版式/编码异常）")

# ========= 广州“东莞目的仓”安全处理：新增字段标签法 =============

def _gbk_find_once(buf: bytes, text: str):
    try:
        patt = text.encode("gbk", errors="ignore")
    except Exception:
        return None
    m = re.search(re.escape(patt), buf)
    return (m.start(), m.end()) if m else None

def find_depot_field_window(buf: bytes, labels=("目的库房", "目的仓"), max_seek=256):
    for lab in labels:
        hit = _gbk_find_once(buf, lab)
        if not hit:
            continue
        _, lab_end = hit
        region = buf[lab_end: lab_end + max_seek]

        i = 0
        while i < len(region) and region[i] in (0x20, 0x09, 0x0D, 0x0A):
            i += 1

        j = i
        while j < len(region):
            b = region[j]
            if b in (0x0D, 0x0A):
                break
            if j - i > 0 and j + 4 < len(region):
                ahead = region[j:j+4]
                if any(kw.encode("gbk", errors="ignore") in ahead for kw in ("目的地","供应商","采购","箱唛","序号")):
                    break
            j += 1

        if j - i >= 8:
            start = lab_end + i
            end = lab_end + j
            return start, end

    return None, None

def patch_pld_with_entry(pld_path: Path, out_path: Path, entry: dict):
    data = pld_path.read_bytes()
    buf = bytearray(data)
    changed, warns, debug = [], [], []

    # 采购单号（*num* / 标签附近 / 兜底）
    star_cnt, star_pos = replace_star_number_all(buf, entry["po"])
    if star_cnt:
        changed.append(f"可视采购单号（*num*）×{star_cnt}")
        debug.append(f"star-number positions: {star_pos}")
    else:
        debug.append("no star-number pattern found")

    pos_list = search_label_positions(buf, "采购单号")
    po_bytes = entry["po"].encode("ascii", errors="ignore")
    label_hits = 0
    for pos in pos_list:
        pairs = find_next_digits_after(buf, pos, max_seek=3000, min_len=6, max_len=20, find_all=True)
        for where, old in pairs:
            replace_region_bytes(buf, where, len(old), po_bytes)
            label_hits += 1
    if label_hits: changed.append(f"采购单号(标签附近)×{label_hits}")
    else: debug.append("no digits found after '采购单号'")

    if (po_bytes not in buf) and (b"*" + po_bytes + b"*" not in buf):
        regions = find_all_digits_regions(buf, min_len=6, max_len=20)
        if regions:
            regions.sort(key=lambda m: len(m.group(1)), reverse=True)
            for m in regions[:2]:
                where = m.start(1); old = m.group(1)
                replace_region_bytes(buf, where, len(old), po_bytes)
            changed.append("采购单号(兜底最长数字)×{}".format(min(2, len(regions))))
        else:
            warns.append("未找到可替换的采购单号区域（请检查模板）")

    # 商家名称
    pos_list = search_label_positions(buf, "商家名称")
    sup_bytes = entry["supplier"].encode("ascii", errors="ignore")
    sup_hits = 0
    for pos in pos_list:
        where, old = find_ascii_word_after(buf, pos, max_seek=3000)
        if where is not None:
            replace_region_bytes(buf, where, len(old), sup_bytes)
            sup_hits += 1
    if sup_hits: changed.append(f"商家名称×{sup_hits}")
    else:
        m = re.search(rb"(?<![A-Za-z0-9])stsnb(?![A-Za-z0-9])", buf)
        if m:
            where = m.start(); old = m.group(0)
            replace_region_bytes(buf, where, len(old), sup_bytes)
            changed.append(f"商家名称(兜底)@{where}")
        else:
            warns.append("未找到可替换的商家名称区域")

    # 箱唛序号（可选）
    if entry.get("no"):
        pos_list = search_label_positions(buf, "箱唛序号")
        no_bytes = str(entry["no"]).encode("ascii", errors="ignore")
        no_hits = 0
        for pos in pos_list:
            pairs = find_next_digits_after(buf, pos, max_seek=3000, min_len=1, max_len=6, find_all=True)
            for where, old in pairs:
                replace_region_bytes(buf, where, len(old), no_bytes)
                no_hits += 1
        if no_hits: changed.append(f"箱唛序号×{no_hits}")
        else: warns.append("未找到可替换的箱唛序号区域")

    # 目的库房（广州+东莞 优先字段标签法；其它情况 先字段法后城市锚点）
    depot_written = False

    if entry["city"] == "广州" and ("东莞" in entry["depot"]):
        fw_start, fw_end = find_depot_field_window(buf, labels=("目的库房", "目的仓"), max_seek=256)
        if fw_start is not None and fw_end is not None and fw_end > fw_start:
            old_len = fw_end - fw_start
            depot_bytes = encode_gbk_fit(entry["depot"], old_len)
            buf[fw_start:fw_end] = depot_bytes
            changed.append(f"目的库房(字段标签法)@{fw_start}-{fw_end}")
            depot_written = True
        else:
            start, end = find_depot_region(buf, entry["city"], prefer_suffix=b"-CHN", window=1024)
            if start is not None and end is not None and end > start:
                old_len = end - start
                depot_bytes = encode_gbk_fit(entry["depot"], old_len)
                buf[start:end] = depot_bytes
                changed.append(f"目的库房(兜底-城市锚点)@{start}-{end}")
                depot_written = True
            else:
                warns.append("广州特例：未定位到目的库房（字段/城市锚点均未命中）")

    if not depot_written:
        fw_start, fw_end = find_depot_field_window(buf, labels=("目的库房", "目的仓"), max_seek=256)
        if fw_start is not None and fw_end is not None and fw_end > fw_start:
            old_len = fw_end - fw_start
            depot_bytes = encode_gbk_fit(entry["depot"], old_len)
            buf[fw_start:fw_end] = depot_bytes
            changed.append(f"目的库房(字段标签法)@{fw_start}-{fw_end}")
            depot_written = True
        else:
            start, end = find_depot_region(buf, entry["city"], prefer_suffix=b"-CHN", window=1024)
            if start is not None and end is not None and end > start:
                old_len = end - start
                depot_bytes = encode_gbk_fit(entry["depot"], old_len)
                buf[start:end] = depot_bytes
                changed.append(f"目的库房(城市锚点)@{start}-{end}")
                depot_written = True
            else:
                warns.append("未定位到目的库房（字段/城市锚点均未命中）")

    # 目的城市显示（省/市/合并 自适应）
    patch_city_display(buf, entry["city"], changed, warns, debug)

    if (po_bytes not in buf) and (b"*" + po_bytes + b"*" not in buf):
        warns.append("改写后未检测到新采购单号明文（条码对象可能以非明文保存）")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_bytes(buf)
    return changed, warns, debug

# ===================== 主流程：一次识别 → 同时产出 标签 + 箱唛 =====================

def main():
    root_dir = Path(__file__).resolve().parent
    # 统一日志目录
    LOG_DIR = root_dir / "日志"
    LOG_DIR.mkdir(parents=True, exist_ok=True)

    # 选择工作簿（同一个）
    try:
        workbook_path = choose_workbook(root_dir)
    except Exception as e:
        if TK_OK: messagebox.showerror("操作已取消", str(e))
        else: print("操作已取消：", e)
        return

    # 打开
    try:
        wb = load_workbook(filename=str(workbook_path), data_only=True)
    except Exception as e:
        if TK_OK: messagebox.showerror("读取失败", f"无法打开工作簿：{workbook_path}\n\n{e}")
        else: print("读取失败：", e)
        return

    # ====== 统一识别（供 标签 + 箱唛 共用）======
    sheet_name_map = resolve_sheet_names(wb)
    b1_values = {exp: (read_b1(wb, real) if real else "") for exp, real in sheet_name_map.items()}
    label_type_simple = decide_label_type_by_b1(b1_values, sheet_name_map)    # "3C" / "玩具"
    label_type_full   = "3C标签" if label_type_simple == "3C" else "玩具标签"
    box_kind_dirname  = "3C箱唛" if label_type_simple == "3C" else "玩具箱唛"

    mmdd = extract_date_tag_from_wb(wb, sheet_name_map)  # MMDD 来自标签 D 列 SN

    # ====== A. 生成 标签 ======
    source_base, tried_paths = resolve_label_template_dir(root_dir, label_type_full)
    if not source_base:
        msg = "未找到标签模板目录：{}\n已尝试：\n{}".format(label_type_full, "\n".join(tried_paths))
        if TK_OK: tk.Tk().withdraw(); messagebox.showerror("未找到模板目录", msg)
        print(msg); return

    out_root_label = root_dir / f"{mmdd}-{label_type_full}"
    out_root_label.mkdir(parents=True, exist_ok=True)

    pld_index, pld_entries = build_pld_index(source_base)
    sheet_to_outfolder = {
        "外仓库配货表": "外星人",
        "梨配货表": "三只梨",
        "兽仓库配货表": "兽",
        "兽无人机拆1": "兽无人机",
    }

    total_copied = 0
    missing_map = {k: [] for k in sheet_to_outfolder.keys()}
    copied_map  = {k: 0  for k in sheet_to_outfolder.keys()}
    detail_lines = []
    soldout_hits = []

    for exp_sheet, outfolder in sheet_to_outfolder.items():
        real_sheet = sheet_name_map.get(exp_sheet)
        if not real_sheet:
            detail_lines.append(f"[错误] 未找到工作表：{exp_sheet}（注意命名差异）")
            continue

        try:
            rows = read_id_sku_e_from_sheet(wb, real_sheet_name=real_sheet, id_col=5, sku_col=1, e_col=5, start_row=2)
        except Exception as e:
            detail_lines.append(f"[错误] 读取工作表 {real_sheet} 失败：{e}")
            continue

        dest_dir = out_root_label / outfolder
        created_dest = False
        count_before = total_copied

        for raw_id, sku, e_val in rows:
            raw_id_str = str(raw_id)
            sku_str = str(sku)
            e_txt = str(e_val)
            use_e_only = (exp_sheet in FORCE_USE_E_ONLY_SHEETS)

            # 清洗“售止”
            if "售止" in raw_id_str:
                cleaned = raw_id_str.replace("售止","")
                soldout_hits.append((exp_sheet, raw_id_str, cleaned))
                raw_id_str = cleaned

            # === 螺旋桨特判（增强版） ===
            forced_name = None
            if e_txt and ("螺旋桨" in e_txt):
                # 使用增强的螺旋桨模板查找函数
                forced_name = find_propeller_template(sku_str, raw_id_str, exp_sheet, pld_index, pld_entries)

            if forced_name:
                cand_names = candidate_filenames(exp_sheet, [], forced_names=[forced_name])
            else:
                id_variants = build_variants(raw_id_str)
                cand_names = candidate_filenames(exp_sheet, id_variants, forced_names=None)

            found = find_matching_templates(pld_index, cand_names)

            if not found and not use_e_only:
                # 数字前缀兜底（对 E-only 表禁用）
                first_nonempty = raw_id_str
                num_pref = numeric_prefix(first_nonempty)
                if num_pref:
                    num_hits = fallback_by_number_prefix(pld_entries, num_pref, exp_sheet)
                    if num_hits:
                        found.extend(num_hits)
                        detail_lines.append(f"[数字前缀兜底] [{exp_sheet}] {raw_id} -> 命中 {len(num_hits)} 个")

            if not found:
                preview = ", ".join(cand_names[:5]) + (f" …共{len(cand_names)}项" if len(cand_names) > 5 else "")
                mark = f"{raw_id}（候选：{preview}）"
                if forced_name:
                    mark += f" ← 螺旋桨映射未命中：{forced_name}"
                missing_map[exp_sheet].append(mark)
                continue

            if not created_dest and not dest_dir.exists():
                dest_dir.mkdir(parents=True, exist_ok=True); created_dest = True

            for src in found:
                try:
                    dst = dest_dir / src.name
                    shutil.copy2(src, dst)
                    total_copied += 1
                except Exception as e:
                    detail_lines.append(f"[复制失败] {src.name} -> {dest_dir}：{e}")

        copied_map[exp_sheet] = total_copied - count_before

    # 统一日志目录
    log_path_label = (root_dir / "日志") / "标签拷贝日志.txt"
    now_str = dt.now().strftime("%Y-%m-%d %H:%M:%S")
    try:
        with log_path_label.open("w", encoding="utf-8") as f:
            f.write(f"脚本版本：{VERSION}\n")
            f.write(f"执行时间：{now_str}\n")
            f.write(f"工作簿：{workbook_path}\n")
            f.write(f"SN提取日期（MMDD）：{mmdd}\n")
            f.write(f"SN日期来源：{DATE_SRC_NOTE}\n")
            f.write(f"判定标签类型：{label_type_full}（模板来源：{source_base}）\n")
            f.write("="*40 + "\n")
            for sheet, folder in sheet_to_outfolder.items():
                f.write(f"【{sheet} → {folder}】复制文件数：{copied_map[sheet]}\n")
                if missing_map[sheet]:
                    f.write("  未找到的编号/候选（候选仅展示前5条）：\n")
                    for m in missing_map[sheet]:
                        f.write(f"    - {m}\n")
            if soldout_hits:
                f.write("\n含“售止”的编号已自动清洗：\n")
                for s, raw_id, cleaned in soldout_hits:
                    f.write(f"  [{s}] {raw_id} -> {cleaned}\n")
            if detail_lines:
                f.write("\n详细信息/兜底记录/错误：\n")
                for line in detail_lines:
                    f.write(line + "\n")
    except Exception:
        pass

    # 标签 .pld 日期批改（报告也写入 LOG_DIR）
    patch_summary, patch_report = run_patch_step(out_root_label, mmdd, ext=".pld", dry=False, make_backup=False, report_dir=(root_dir / "日志"))

    # ====== B. 生成 箱唛 ======
    ws_box = find_box_sheet(wb)
    store_sub = decide_store_subfolder(ws_box)
    entries = parse_entries(ws_box)
    if not entries:
        msg = "未在“箱唛”工作表中识别到任何城市条目（需包含 目的地/供应商简码/采购单号/目的库房）。"
        if TK_OK: messagebox.showwarning("未识别", msg)
        print(msg)
        lines = [
            f"复制完成（标签），共复制 {total_copied} 个文件。",
            f"标签类型：{label_type_full} | SN日期(MMDD)：{mmdd}",
            f"标签模板目录：{source_base}",
            f"标签输出目录：{out_root_label}",
            f"标签拷贝日志：{log_path_label}",
            "",
            "【标签第二步：批量修改 .pld 日期】",
            patch_summary,
            f"标签日期报告：{patch_report}",
        ]
        print("\n".join(lines))
        if TK_OK:
            try:
                tk.Tk().withdraw()
                messagebox.showinfo("完成（仅标签）", "\n".join(lines))
            except Exception:
                pass
        return

    box_tpl_dir, tried_box = resolve_box_template_dir(root_dir, box_kind_dirname)
    if not box_tpl_dir:
        msg = f"未找到箱唛模板目录：{box_kind_dirname}\n已尝试：\n" + "\n".join(tried_box)
        if TK_OK: messagebox.showerror("未找到模板", msg)
        print(msg); return

    out_root_box = root_dir / f"{mmdd}-{box_kind_dirname}"
    if store_sub:
        out_root_box = out_root_box / store_sub
    out_root_box.mkdir(parents=True, exist_ok=True)

    city_counts = {}
    summary_lines, debug_lines = [], []
    total_ok = total_warn = 0

    for ent in entries:
        city = ent["city"]
        tpl = find_city_template(box_tpl_dir, city)
        if not tpl:
            msg = f"[跳过] {city}：未找到对应模板（{box_tpl_dir}）"
            summary_lines.append(msg); total_warn += 1
            continue

        cnt = city_counts.get(city, 0) + 1
        city_counts[city] = cnt
        out_path = out_root_box / (tpl.name if cnt == 1 else f"{tpl.stem}-{cnt}{tpl.suffix}")

        shutil.copy2(tpl, out_path)
        changed, warns, debug = patch_pld_with_entry(out_path, out_path, ent)

        if warns:
            total_warn += 1
            summary_lines.append(f"[{city}] {out_path.name}\n  修改：{'; '.join(changed) if changed else '无'}\n  提示：{'; '.join(warns)}")
        else:
            total_ok += 1
            summary_lines.append(f"[{city}] {out_path.name}\n  修改：{'; '.join(changed) if changed else '无'}")

        if debug:
            debug_lines.append(f"[{city}] {out_path.name} -> " + " | ".join(debug))

    # 箱唛日志 -> LOG_DIR
    log_path_box = (root_dir / "日志") / "箱唛处理日志.txt"
    head = [
        f"脚本版本：{VERSION}",
        f"执行时间：{dt.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"工作簿：{workbook_path}",
        f"类型：{label_type_simple}  |  模板目录：{box_tpl_dir}",
        f"输出目录：{out_root_box}",
        f"生成成功：{total_ok}  |  有提示/检查：{total_warn}",
        "-" * 60,
    ]
    with log_path_box.open("w", encoding="utf-8") as f:
        f.write("\n".join(head) + "\n")
        for line in summary_lines:
            f.write(line + "\n")
        if debug_lines:
            f.write("\n[调试]\n")
            for d in debug_lines:
                f.write(d + "\n")

    # ====== 汇总弹窗 / 控制台输出 ======
    lines = [
        f"复制完成（标签），共复制 {total_copied} 个文件。",
        f"标签类型：{label_type_full} | SN日期(MMDD)：{mmdd}",
        f"标签模板目录：{source_base}",
        f"标签输出目录：{out_root_label}",
        f"标签拷贝日志：{log_path_label}",
        "",
        "【标签 .pld 日期批改（自动写回）】",
        patch_summary,
        f"标签日期报告：{(root_dir / '日志') / 'patch_pld_report.txt'}",
        "",
        f"箱唛模板目录：{box_tpl_dir}",
        f"箱唛输出目录：{out_root_box}",
        f"箱唛生成成功：{total_ok}  |  有提示/检查：{total_warn}",
        f"箱唛日志：{log_path_box}",
    ]
    print("\n".join(lines))
    if TK_OK:
        try:
            tk.Tk().withdraw()
            messagebox.showinfo("完成（标签 + 箱唛）", "\n".join(lines))
        except Exception:
            pass

# ===================== 仅预定表功能（新增）=====================

def read_reservation_data_from_sheet(wb, real_sheet_name: str, sku_col: int = 1, city_start_col: int = 7, city_end_col: int = 14, start_row: int = 2):
    """
    读取配货表中的商品编号(A列)和城市列数据
    
    参数:
        wb: 工作簿对象
        real_sheet_name: 工作表名称
        sku_col: 商品编号列（默认A列=1）
        city_start_col: 城市列起始（默认G列=7）
        city_end_col: 城市列结束（默认N列=14，会自动扩展到最后一列）
        start_row: 起始行（默认第2行）
    
    返回:
        (city_names, data_rows)
        city_names: 城市名称列表
        data_rows: [(商品编号, [城市1数量, 城市2数量, ...]), ...]
    """
    ws = wb[real_sheet_name]
    
    # 定义标准城市列表（用于识别城市列）
    standard_cities = ["北京", "上海", "广州", "成都", "武汉", "沈阳", "西安", "德州"]
    
    # 自动检测城市列：从 city_start_col 开始，找到所有包含城市名或城市+数字的列
    city_names = []
    city_cols = []
    
    # 导入正则表达式
    import re
    
    # 从 city_start_col 开始识别城市列，直到遇到空单元格或非城市格式的单元格时停止
    # 这样可以避免读取表右边的其他表
    for col in range(city_start_col, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        
        # 如果单元格为空，停止识别
        if not cell_value:
            break
        
        city_name = str(cell_value).strip()
        
        # 检查是否是城市名或城市+数字
        is_city = False
        if city_name in standard_cities:
            is_city = True
        else:
            # 检查是否是城市+数字的格式（必须完全匹配）
            for std_city in standard_cities:
                if re.match(f"^{std_city}[-]?\\d+$", city_name):
                    is_city = True
                    break
        
        if is_city:
            city_names.append(city_name)
            city_cols.append(col)
        else:
            # 遇到非城市格式的单元格，停止识别
            break
    
    # 调试：打印读取的城市名称和列号
    import sys
    print(f"[DEBUG] 读取的城市名称：{city_names}", file=sys.stderr)
    print(f"[DEBUG] 对应的列号：{city_cols}", file=sys.stderr)
    
    # 打印前几行的原始数据用于对比
    print(f"[DEBUG] 表头行（第1行）：", file=sys.stderr)
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        print(f"  列{col}({chr(64+col)}): {val}", file=sys.stderr)
    
    # 读取数据行
    data_rows = []
    seen_skus = set()
    
    for row in ws.iter_rows(min_row=start_row):
        # 读取商品编号（A列）
        sku_cell = row[sku_col - 1]
        sku_value = sku_cell.value
        
        if sku_value is None:
            continue
        
        sku = str(sku_value).strip()
        if not sku or sku in seen_skus:
            continue
        
        # 只保留纯数字的商品编号，过滤掉"合计"、"未购单号"等文字
        if not sku.isdigit():
            continue
        
        seen_skus.add(sku)
        
        # 读取城市列的数据（使用检测到的 city_cols）
        city_quantities = []
        for col in city_cols:
            cell = row[col - 1]
            value = cell.value
            
            # 转换为数字，如果不是数字则为0
            try:
                qty = int(value) if value is not None else 0
            except (ValueError, TypeError):
                qty = 0
            
            city_quantities.append(qty)
        
        data_rows.append((sku, city_quantities))
    
    return city_names, data_rows


def generate_reservation_table(wb, sheet_name_map: dict, output_path: Path, callback=None):
    """
    生成仅预定表：转置商品编号和城市数据
    
    参数:
        wb: 工作簿对象
        sheet_name_map: 工作表名称映射
        output_path: 输出Excel文件路径
        callback: 日志回调函数
    
    返回:
        统计信息字典
    """
    def log(msg):
        if callback:
            callback(msg)
    
    from openpyxl import Workbook
    
    # 需要处理的3C表单（按顺序）
    target_sheets = ["外仓库配货表", "梨配货表", "兽仓库配货表", "兽无人机仓库配货表总"]
    
    # 定义城市顺序（与G-N列对应）
    city_order = ["北京", "上海", "广州", "成都", "武汉", "沈阳", "西安", "德州"]
    
    # 定义城市别名映射（将特殊城市映射为标准城市，并记录原始名称用于备注）
    # 支持"城市+数字"或"城市-数字"的格式，自动映射为标准城市
    city_alias_map = {
        "北京2": ("北京", "北京2"),
        "北京-2": ("北京", "北京2"),
        "上海2": ("上海", "上海2"),
        "上海-2": ("上海", "上海2"),
        "广州2": ("广州", "广州2"),
        "广州-2": ("广州", "广州2"),
        "成都2": ("成都", "成都2"),
        "成都-2": ("成都", "成都2"),
        "武汉2": ("武汉", "武汉2"),
        "武汉-2": ("武汉", "武汉2"),
        "沈阳2": ("沈阳", "沈阳2"),
        "沈阳-2": ("沈阳", "沈阳2"),
        "西安2": ("西安", "西安2"),
        "西安-2": ("西安", "西安2"),
        "德州2": ("德州", "德州2"),
        "德州-2": ("德州", "德州2"),
    }
    
    # 导入正则表达式模块
    import re
    
    # 收集所有数据，记录工作表来源和城市
    all_data = []  # [(工作表序号, 城市序号, 商品编号, 配送中心名称, 数量, 备注), ...]
    
    for sheet_idx, exp_sheet in enumerate(target_sheets):
        real_sheet = sheet_name_map.get(exp_sheet)
        if not real_sheet:
            log(f"  跳过：{exp_sheet}（未找到工作表）")
            continue
        
        try:
            log(f"  读取工作表：{real_sheet}")
            city_names, data_rows = read_reservation_data_from_sheet(wb, real_sheet)
            
            # 转置数据
            for sku, quantities in data_rows:
                for i, qty in enumerate(quantities):
                    if qty > 0:  # 只记录数量大于0的
                        city_name = city_names[i]
                        # 检查是否需要映射城市名
                        remark = ""
                        mapped_city = city_name
                        
                        # 首先检查精确匹配
                        if city_name in city_alias_map:
                            mapped_city, remark = city_alias_map[city_name]
                        else:
                            # 尝试正则匹配：城市名+数字 或 城市名-数字
                            for standard_city in city_order:
                                # 匹配"城市+数字"或"城市-数字"
                                pattern = f"^{standard_city}[-]?\\d+$"
                                if re.match(pattern, city_name):
                                    mapped_city = standard_city
                                    remark = city_name  # 备注为原始名称
                                    break
                        
                        # 获取城市在标准顺序中的序号
                        try:
                            city_idx = city_order.index(mapped_city)
                        except ValueError:
                            city_idx = 999  # 未知城市排在最后
                        all_data.append((sheet_idx, city_idx, sku, mapped_city, qty, remark))
            
            log(f"    提取 {len(data_rows)} 个商品编号")
        except Exception as e:
            log(f"  ✗ 读取失败：{e}")
            import traceback
            log(traceback.format_exc())
            continue
    
    if not all_data:
        log("✗ 未提取到任何预定数据")
        return {"success": False, "total_rows": 0}
    
    # 排序：先按工作表序号，再按城市序号，最后按商品编号
    all_data.sort(key=lambda x: (x[0], x[1], x[2]))
    
    log(f"  数据已按工作表和配送中心排序")
    
    # 创建新的工作簿
    new_wb = Workbook()
    ws = new_wb.active
    ws.title = "预定表"
    
    # 导入样式模块
    from openpyxl.styles import Alignment
    
    # 写入表头
    ws['A1'] = "商品编号"
    ws['B1'] = "配送中心名称"
    ws['C1'] = "有限预订数量"
    ws['D1'] = "备注"
    
    # 设置表头居中
    for cell in ['A1', 'B1', 'C1', 'D1']:
        ws[cell].alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入数据（提取实际需要的字段：商品编号、城市、数量、备注）
    for idx, (sheet_idx, city_idx, sku, city, qty, remark) in enumerate(all_data, start=2):
        # 商品编号保持为文本格式（与原表一致）
        ws[f'A{idx}'] = sku
        
        ws[f'B{idx}'] = city
        ws[f'C{idx}'] = qty
        ws[f'D{idx}'] = remark
        
        # 设置每行数据居中
        ws[f'A{idx}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'B{idx}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'C{idx}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'D{idx}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # 自动调整列宽
    from openpyxl.utils import get_column_letter
    
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        
        for cell in column_cells:
            try:
                if cell.value:
                    # 计算内容长度（中文字符按2个字符计算）
                    cell_value = str(cell.value)
                    length = sum(2 if ord(c) > 127 else 1 for c in cell_value)
                    max_length = max(max_length, length)
            except:
                pass
        
        # 设置列宽（加一点边距）
        adjusted_width = min(max_length + 2, 50)  # 最大宽度限制为50
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 保存文件
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # 检查文件是否已存在，如果存在则创建副本（避免文件被占用导致保存失败）
    if output_path.exists():
        import time
        timestamp = time.strftime("%H%M%S")
        base_name = output_path.stem
        new_name = f"{base_name}_{timestamp}{output_path.suffix}"
        output_path = output_path.parent / new_name
        log(f"  文件已存在，创建副本：{output_path}")
    
    try:
        new_wb.save(str(output_path))
    except PermissionError:
        # 如果仍然因权限问题无法保存，再创建副本
        import time
        timestamp = time.strftime("%H%M%S%f")
        base_name = output_path.stem
        new_name = f"{base_name}_{timestamp}{output_path.suffix}"
        output_path = output_path.parent / new_name
        log(f"  保存失败，创建新副本：{output_path}")
        new_wb.save(str(output_path))
    
    log(f"✓ 预定表生成完成：{output_path}")
    log(f"  共 {len(all_data)} 行数据")
    
    return {
        "success": True,
        "total_rows": len(all_data),
        "output_path": str(output_path)
    }


if __name__ == "__main__":
    try:
        main()
    except Exception:
        err = traceback.format_exc()
        if TK_OK:
            try:
                tk.Tk().withdraw()
                messagebox.showerror("程序异常", err)
            except Exception:
                pass
        print(err)
