"""
标签箱唛处理包装器（整理后版本）
适用于新的文件结构
"""
import sys
from pathlib import Path
from openpyxl import load_workbook

# 导入核心模块
import importlib.util
spec = importlib.util.spec_from_file_location("label_box_core", Path(__file__).parent / "core.py")
core = importlib.util.module_from_spec(spec)
sys.modules['label_box_core'] = core
spec.loader.exec_module(core)


def read_green_rows_from_sheet(wb, real_sheet_name, id_col=5, sku_col=1, e_col=5, start_row=2, debug=False):
    """
    读取工作表中编号不是红色文字的行（排除红字，其他都要）
    
    参数:
        wb: 工作簿对象
        real_sheet_name: 工作表名称
        id_col: 编号列（从1开始，默认E列=5）
        sku_col: SKU列（从1开始，默认A列=1）
        e_col: E列（从1开始，默认E列=5）
        start_row: 起始行（默认第2行）
        debug: 是否输出调试信息
    
    返回:
        [(id, sku, e_val), ...] 非红字的行列表
        如果debug=True，还返回调试信息字典
    """
    ws = wb[real_sheet_name]
    rows = []
    seen = set()
    debug_info = {'total_rows': 0, 'green_count': 0, 'other_count': 0, 'colors': [], 'method': '排除红字'}
    
    for row_num, r in enumerate(ws.iter_rows(min_row=start_row), start=start_row):
        # 获取编号单元格（E列）
        id_cell = r[id_col - 1]
        idv = id_cell.value
        skuv = r[sku_col - 1].value
        e_val = r[e_col - 1].value
        
        if idv is None and skuv is None and e_val is None:
            continue
        
        debug_info['total_rows'] += 1
        
        # 简化逻辑：红字不要，其他都要
        is_red = False
        color_desc = "未检测"
        
        # 检查字体颜色是否为红色
        if id_cell.font and id_cell.font.color:
            font_color = id_cell.font.color
            try:
                if hasattr(font_color, 'rgb') and font_color.rgb is not None:
                    # 尝试转换RGB值
                    if isinstance(font_color.rgb, str):
                        font_rgb = font_color.rgb.upper()
                    else:
                        font_rgb = str(font_color.rgb).upper()
                    
                    # 检测红色：FFFF0000或FF0000
                    if 'FF0000' in font_rgb:
                        is_red = True
                        color_desc = f"红字({font_rgb})"
                    else:
                        color_desc = f"非红字({font_rgb})"
            except:
                # 无法读取颜色，当作非红字处理
                color_desc = "其他字色"
        else:
            # 没有字体颜色设置，当作默认黑色，保留
            color_desc = "默认字色"
        
        if debug:
            debug_info['colors'].append({'row': r[0].row, 'id': idv, 'color': color_desc, 'is_red': is_red})
        
        # 排除红字的行
        if is_red:
            debug_info['other_count'] += 1
            continue
        
        debug_info['green_count'] += 1
        
        idstr = str(idv).strip() if idv is not None else ""
        sku = str(skuv).strip() if skuv is not None else ""
        
        if not idstr and not sku:
            continue
        
        # 过滤：只保留 A 列（SKU）是数字的行
        sku_is_number = sku and sku.isdigit()
        if not sku_is_number:
            continue
        
        k = (idstr, sku)
        if k in seen:
            continue
        seen.add(k)
        
        rows.append((idstr, sku, e_val, row_num))
    
    if debug:
        return rows, debug_info
    return rows, None


def process_excel_file(workbook_path, output_base=None, callback=None, progress_callback=None,
                       type_mode="auto", output_mode="both", create_zip=False, save_log=False, selected_shops=None, label_type_simple_override=None):
    """
    处理Excel文件的包装函数（整理后版本）
    
    参数:
        workbook_path: Excel文件路径
        output_base: 输出目录路径（可选，默认为系统下载文件夹）
        callback: 回调函数，用于输出日志信息
        progress_callback: 进度回调函数 progress_callback(value, text)
        type_mode: 类型模式 ("auto"/"3c"/"toy")
        output_mode: 输出模式 ("both"/"label"/"box")
        create_zip: 是否创建ZIP
        save_log: 是否保存日志文件
        selected_shops: 选中的店铺列表（用于标签筛选）
    
    返回:
        处理结果字典
    """
    def log(msg):
        if callback:
            callback(msg)
        else:
            print(msg)
    
    def progress(value, text=None):
        if progress_callback:
            progress_callback(value, text)
    
    # 获取项目根目录（兼容开发环境和打包后环境）
    if getattr(sys, 'frozen', False):
        # 打包后：PyInstaller把数据文件放在_internal，但templates在_internal下
        # 所以root_dir应该指向_internal
        root_dir = Path(sys.executable).parent / "_internal"
    else:
        # 开发环境：向上3级到Gabrix根目录
        root_dir = Path(__file__).resolve().parent.parent.parent.parent
    
    # 定义路径
    template_base = root_dir / "templates"      # 模板根目录
    
    # 如果没有指定输出路径，使用系统下载文件夹
    if output_base is None:
        import os
        output_base = Path(os.path.expanduser("~")) / "Downloads"
    else:
        output_base = Path(output_base)
    
    # 确保输出目录存在
    output_base.mkdir(parents=True, exist_ok=True)
    
    workbook_path = Path(workbook_path)
    
    try:
        progress(10, "正在打开工作簿...")
        log(f"正在打开工作簿：{workbook_path.name}...")
        wb = load_workbook(filename=str(workbook_path), data_only=True)
    except Exception as e:
        log(f"✗ 无法打开工作簿：{e}")
        return {"success": False, "error": str(e)}
    
    try:
        # 统一识别
        progress(15, "正在识别工作表...")
        log("正在识别工作表...")
        sheet_name_map = core.resolve_sheet_names(wb)
        b1_values = {exp: (core.read_b1(wb, real) if real else "") 
                     for exp, real in sheet_name_map.items()}
        
        # 先自动识别
        auto_detected_type = core.decide_label_type_by_b1(b1_values, sheet_name_map)
        
        # 根据模式决定类型
        if type_mode == "3c":
            label_type_simple = "3C"
            log(f"类型：强制3C模式（自动识别为：{auto_detected_type}）")
        elif type_mode == "toy":
            label_type_simple = "玩具"
            log(f"类型：强制玩具模式（自动识别为：{auto_detected_type}）")
        else:
            label_type_simple = auto_detected_type
            log(f"类型：自动识别为 {label_type_simple}")
        
        label_type_full = "3C标签" if label_type_simple == "3C" else "玩具标签"
        label_type_name = "3C" if label_type_simple == "3C" else "玩具"
        box_kind_dirname = "3C箱唛" if label_type_simple == "3C" else "玩具箱唛"
        
        log(f"识别类型：{label_type_full}")
        
        progress(20, "提取日期信息...")
        mmdd = core.extract_date_tag_from_wb(wb, sheet_name_map)
        log(f"提取日期(MMDD)：{mmdd}")
        log(f"日期来源：{core.DATE_SRC_NOTE}")
        
        # 不再创建统一的主文件夹，标签和箱唛分别输出
        log(f"\n输出基础目录：{output_base}")
        
        # 特殊处理：仅预定表模式（只针对3C）
        if output_mode == "reservation":
            if label_type_simple != "3C":
                msg = "仅预定表功能只适用于3C配货表"
                log(f"✗ {msg}")
                return {"success": False, "error": msg}
            
            progress(25, "正在生成预定表...")
            log("\n=== 生成仅预定表 ===")
            
            # 输出文件：1028-3C预定表.xlsx
            output_filename = f"{mmdd}-3C预定表.xlsx"
            output_path = output_base / output_filename
            
            result = core.generate_reservation_table(wb, sheet_name_map, output_path, callback=callback)
            
            if result["success"]:
                progress(100, "✓ 预定表生成完成！")
                log(f"\n✓ 预定表生成完成：{output_path}")
                return {
                    "success": True,
                    "reservation_only": True,
                    "total_rows": result["total_rows"],
                    "output_path": str(output_path),
                    "main_output": str(output_base),
                    "mmdd": mmdd,
                    "label_type": label_type_full,
                    "label_type_name": label_type_name
                }
            else:
                return {"success": False, "error": "生成预定表失败"}
        
        # A. 生成标签 (如果需要)
        total_copied = 0
        out_root_label = None
        
        if output_mode in ["both", "label"]:
            progress(25, "开始生成标签...")
            log("\n=== 开始生成标签 ===")
            
            # 使用新的模板路径
            source_base = template_base / "标签模板" / label_type_full
            if not source_base.exists():
                msg = f"未找到标签模板目录：{source_base}"
                log(f"✗ {msg}")
                return {"success": False, "error": msg}
            
            log(f"标签模板目录：{source_base}")
            
            # 标签输出到独立文件夹：1028-3C标签
            out_root_label = output_base / f"{mmdd}-{label_type_name}标签"
            out_root_label.mkdir(parents=True, exist_ok=True)
            log(f"标签输出目录：{out_root_label}")
            
            pld_index, pld_entries = core.build_pld_index(source_base)
            log(f"已索引 {len(pld_index)} 个标签模板文件")
            
            # 根据类型选择不同的文件夹名称
            if label_type_simple == "3C":
                sheet_to_outfolder = {
                    "外仓库配货表": "外星人玩具",
                    "梨配货表": "三只梨",
                    "兽仓库配货表": "兽",
                    "兽无人机拆1": "兽无人机拆1",
                    "兽无人机拆2": "兽无人机拆2",
                }
            else:  # 玩具
                sheet_to_outfolder = {
                    "外仓库配货表": "外星人",
                    "兽仓库配货表": "兽模型",
                }
            
            # 如果指定了店铺筛选，只处理选中的店铺
            if selected_shops:
                filtered_sheet_to_outfolder = {k: v for k, v in sheet_to_outfolder.items() if v in selected_shops}
                if filtered_sheet_to_outfolder:
                    sheet_to_outfolder = filtered_sheet_to_outfolder
                    log(f"已启用店铺筛选，只处理：{', '.join(selected_shops)}")
            
            missing_map = {k: [] for k in sheet_to_outfolder.keys()}
            missing_details = {k: [] for k in sheet_to_outfolder.keys()}  # 存储缺少标签的详细信息
            copied_map = {k: 0 for k in sheet_to_outfolder.keys()}
            total_expected = 0  # 应该生成的总标签数
            
            progress(30, "正在复制标签文件...")
            for exp_sheet, outfolder in sheet_to_outfolder.items():
                real_sheet = sheet_name_map.get(exp_sheet)
                if not real_sheet:
                    log(f"  跳过：{exp_sheet}（未找到工作表）")
                    continue
                
                log(f"\n处理工作表：{exp_sheet} → {outfolder}")
                
                try:
                    # 特殊处理：兽无人机拆2排除红字行
                    if exp_sheet == "兽无人机拆2":
                        rows, debug_info = read_green_rows_from_sheet(wb, real_sheet_name=real_sheet, debug=True)
                        log(f"  筛选非红字行：{len(rows)} 个")
                        if debug_info:
                            log(f"  调试信息：检测到 {debug_info['total_rows']} 行，保留 {debug_info['green_count']} 行，排除红字 {debug_info['other_count']} 行")
                    else:
                        rows = core.read_id_sku_e_from_sheet(wb, real_sheet_name=real_sheet)
                except Exception as e:
                    log(f"  ✗ 读取失败：{e}")
                    continue
                
                dest_dir = out_root_label / outfolder
                count_before = total_copied
                found_count = 0
                # 统计应该生成的标签数（只计入数字SKU的行，已在read_id_sku_e_from_sheet中过滤）
                total_expected += len(rows)
                log(f"  读取行数：{len(rows)} 个（已过滤非数字SKU）")
                
                for row_data in rows:
                    # 处理新的四元组格式 (id, sku, e_val, row_num)
                    if len(row_data) == 4:
                        raw_id, sku, e_val, row_num = row_data
                    else:
                        raw_id, sku, e_val = row_data
                        row_num = None
                    
                    # 如果 raw_id 为空，使用行号作为标识符
                    if raw_id:
                        raw_id_str = str(raw_id)
                    elif row_num:
                        raw_id_str = f"(第{row_num}行)"
                    else:
                        raw_id_str = "(未知)"
                    
                    e_txt = str(e_val)
                    row_idx = row_num
                    
                    # 螺旋桨特判（增强版）
                    forced_name = None
                    if e_txt and ("螺旋桨" in e_txt):
                        forced_name = core.find_propeller_template(str(sku), raw_id_str, exp_sheet, pld_index, pld_entries)
                        if forced_name:
                            log(f"  [螺旋桨] SKU {sku} → 匹配到：{forced_name}")
                    
                    if forced_name:
                        cand_names = core.candidate_filenames(exp_sheet, [], forced_names=[forced_name])
                    else:
                        id_variants = core.build_variants(raw_id_str)
                        cand_names = core.candidate_filenames(exp_sheet, id_variants)
                    
                    found = core.find_matching_templates(pld_index, cand_names)
                    
                    if found:
                        if not dest_dir.exists():
                            dest_dir.mkdir(parents=True, exist_ok=True)
                        
                        for src in found:
                            import shutil
                            try:
                                dst = dest_dir / src.name
                                shutil.copy2(src, dst)
                                total_copied += 1
                                found_count += 1
                            except Exception as e:
                                log(f"  复制失败：{src.name} → {e}")
                    else:
                        # 标签未找到，添加到缺少列表
                        # 再次检查 SKU 是否是数字，排除非数据行
                        sku_is_number = str(sku).isdigit() if sku else False
                        if sku_is_number:
                            # 优先使用 SKU 作为标识符，如果 SKU 为空则使用 raw_id_str
                            identifier = str(sku) if sku else raw_id_str
                            missing_map[exp_sheet].append(identifier)
                            # 记录缺少标签的详细信息（商品编号和E列值）
                            missing_details[exp_sheet].append({
                                'sku': str(sku) if sku else raw_id_str,
                                'id': raw_id_str,
                                'e_val': e_txt,
                                'row': row_idx
                            })
                            log(f"  [缺少] SKU {sku}：未找到模板（候选：{cand_names}）")
                        else:
                            # 非数字 SKU 的行，不统计为缺少
                            log(f"  [跳过] 非数字SKU行：SKU={sku}, E={e_txt}")
                        if forced_name:
                            log(f"  [警告] 螺旋桨文件未找到：{forced_name}（候选：{cand_names}）")
                
                copied_map[exp_sheet] = total_copied - count_before
                log(f"  ✓ 已复制 {found_count} 个文件")
            
            log(f"\n标签复制完成，共复制 {total_copied} 个文件")
            
            # 检查是否有螺旋桨文件未找到
            propeller_missing = []
            for sheet, missing_ids in missing_map.items():
                if not missing_ids:
                    continue
                try:
                    real_sheet = sheet_name_map.get(sheet)
                    if real_sheet:
                        rows = core.read_id_sku_e_from_sheet(wb, real_sheet_name=real_sheet)
                        for row_data in rows:
                            # 处理四元组格式 (id, sku, e_val, row_num)
                            if len(row_data) == 4:
                                raw_id, sku, e_val, row_num = row_data
                            else:
                                raw_id, sku, e_val = row_data
                                row_num = 0
                            raw_id_str = str(raw_id)
                            # 检查这个ID是否在missing_ids中，且E列包含"螺旋桨"
                            if raw_id_str in missing_ids and e_val and "螺旋桨" in str(e_val):
                                propeller_missing.append({
                                    'sheet': sheet,
                                    'sku': sku,
                                    'id': raw_id_str
                                })
                except Exception as e:
                    log(f"检查螺旋桨未找到时出错：{e}")
            
            # 如果有螺旋桨文件未找到，记录到返回结果中
            if propeller_missing:
                log(f"\n[警告] 发现 {len(propeller_missing)} 个螺旋桨文件未找到：")
                for item in propeller_missing:
                    log(f"  - {item['sheet']} / SKU: {item['sku']} / ID: {item['id']}")
            
            # 标签日期批改
            progress(60, "正在批改标签日期...")
            log("\n=== 批量修改标签日期 ===")
            
            # 不再单独保存patch报告文件，直接合并到日志中
            patch_summary, patch_report = core.run_patch_step(
                out_root_label, mmdd, ext=".pld", 
                dry=False, make_backup=False, report_dir=None
            )
            log(patch_summary)
            
            # 显示完整的批改报告
            if patch_report:
                # 确保是字符串类型
                report_text = str(patch_report) if not isinstance(patch_report, str) else patch_report
                if report_text and len(report_text) > 0:
                    log("\n批改详细报告：")
                    log("-" * 50)
                    lines = report_text.split('\n')
                    for line in lines:
                        if line.strip():
                            log(f"{line}")
                    log("-" * 50)
        
        # B. 生成箱唛 (如果需要)
        if output_mode in ["both", "box"]:
            progress(70, "开始生成箱唛...")
            log("\n=== 开始生成箱唛 ===")
            ws_box = core.find_box_sheet(wb)
        else:
            ws_box = None
        
        if not ws_box:
            log("✗ 未找到箱唛工作表")
            return {
                "success": True,
                "label_only": True,
                "total_copied": total_copied,
                "main_output": str(output_base),
                "label_output": str(out_root_label) if out_root_label else None,
                "mmdd": mmdd,
                "label_type": label_type_full,
                "label_type_name": label_type_name,
                "auto_detected_type": auto_detected_type,
                "used_type": label_type_simple,
                "type_mismatch": (type_mode != "auto" and label_type_simple != auto_detected_type)
            }
        
        store_sub = core.decide_store_subfolder(ws_box)
        entries = core.parse_entries(ws_box)
        
        if not entries:
            log("✗ 箱唛工作表中未识别到有效条目")
            return {
                "success": True,
                "label_only": True,
                "total_copied": total_copied,
                "main_output": str(output_base),
                "label_output": str(out_root_label) if out_root_label else None,
                "mmdd": mmdd,
                "label_type": label_type_full,
                "label_type_name": label_type_name,
                "auto_detected_type": auto_detected_type,
                "used_type": label_type_simple,
                "type_mismatch": (type_mode != "auto" and label_type_simple != auto_detected_type)
            }
        
        log(f"识别到 {len(entries)} 个箱唛条目")
        
        # 使用新的模板路径
        box_tpl_dir = template_base / "箱唛模板" / box_kind_dirname
        if not box_tpl_dir.exists():
            log(f"✗ 未找到箱唛模板目录：{box_tpl_dir}")
            return {
                "success": True,
                "label_only": True,
                "total_copied": total_copied,
                "main_output": str(output_base),
                "label_output": str(out_root_label) if out_root_label else None,
                "mmdd": mmdd,
                "label_type": label_type_full,
                "label_type_name": label_type_name,
                "auto_detected_type": auto_detected_type,
                "used_type": label_type_simple,
                "type_mismatch": (type_mode != "auto" and label_type_simple != auto_detected_type)
            }
        
        log(f"箱唛模板目录：{box_tpl_dir}")
        
        # 箱唛输出到独立文件夹：1028-3C箱唛
        out_root_box = output_base / f"{mmdd}-{label_type_name}箱唛"
        if store_sub:
            out_root_box = out_root_box / store_sub
        out_root_box.mkdir(parents=True, exist_ok=True)
        log(f"箱唛输出目录：{out_root_box}")
        
        city_counts = {}
        total_ok = total_warn = 0
        
        progress(75, "正在生成箱唛文件...")
        for ent in entries:
            city = ent["city"]
            tpl = core.find_city_template(box_tpl_dir, city)
            
            if not tpl:
                log(f"  跳过 {city}：未找到模板")
                total_warn += 1
                continue
            
            cnt = city_counts.get(city, 0) + 1
            city_counts[city] = cnt
            out_path = out_root_box / (tpl.name if cnt == 1 else f"{tpl.stem}-{cnt}{tpl.suffix}")
            
            import shutil
            shutil.copy2(tpl, out_path)
            changed, warns, debug = core.patch_pld_with_entry(out_path, out_path, ent)
            
            if warns:
                total_warn += 1
                log(f"  [{city}] {out_path.name} - 有警告")
            else:
                total_ok += 1
                log(f"  ✓ [{city}] {out_path.name}")
        
        log(f"\n箱唛生成完成：成功 {total_ok} 个，警告 {total_warn} 个")
        
        progress(95, "处理完成...")
        
        # 日志保存在UI层处理
        
        # ZIP打包功能
        if create_zip:
            progress(97, "正在打包ZIP...")
            log("\n=== 打包ZIP文件 ===")
            import shutil
            
            zip_files = []
            try:
                # 打包标签文件夹
                if out_root_label and out_root_label.exists():
                    zip_name = f"{mmdd}-{label_type_name}标签"
                    zip_path = output_base / zip_name
                    shutil.make_archive(str(zip_path), 'zip', str(out_root_label))
                    zip_files.append(f"{zip_name}.zip")
                    log(f"  ✓ 已打包：{zip_name}.zip")
                
                # 打包箱唛文件夹
                if output_mode in ["both", "box"] and out_root_box and out_root_box.exists():
                    zip_name = f"{mmdd}-{label_type_name}箱唛"
                    zip_path = output_base / zip_name
                    shutil.make_archive(str(zip_path), 'zip', str(out_root_box))
                    zip_files.append(f"{zip_name}.zip")
                    log(f"  ✓ 已打包：{zip_name}.zip")
                
                log(f"\nZIP打包完成，共生成 {len(zip_files)} 个压缩包")
            except Exception as zip_err:
                log(f"✗ ZIP打包失败：{zip_err}")
        
        progress(100, "✓ 全部完成！")
        
        return {
            "success": True,
            "total_copied": total_copied,
            "total_expected": total_expected,  # 应该生成的总标签数
            "total_missing": len([id for ids in missing_map.values() for id in ids]),  # 缺少的总标签数
            "missing_map": missing_map,  # 缺少的标签详情
            "main_output": str(output_base),
            "label_output": str(out_root_label) if out_root_label else None,
            "box_output": str(out_root_box) if output_mode in ["both", "box"] else None,
            "box_ok": total_ok,
            "box_warn": total_warn,
            "mmdd": mmdd,
            "label_type": label_type_full,
            "label_type_name": label_type_name,
            "auto_detected_type": auto_detected_type,
            "used_type": label_type_simple,
            "type_mismatch": (type_mode != "auto" and label_type_simple != auto_detected_type),
            "propeller_missing": propeller_missing  # 添加螺旋桨未找到的信息
        }
        
    except Exception as e:
        import traceback
        error_trace = traceback.format_exc()
        log(f"\n✗ 处理异常：{e}")
        log(error_trace)
        return {"success": False, "error": str(e), "traceback": error_trace}


if __name__ == "__main__":
    # 命令行测试
    if len(sys.argv) > 1:
        result = process_excel_file(sys.argv[1])
        print("\n最终结果：", result)
    else:
        print("用法：python 标签箱唛_wrapper_organized.py <Excel文件路径>")

